"""
ByteCredits – NFC attendance & payment. Flask + SQLite backend.
After login, user without a registered card is asked to register card (roll no + card UID via mobile NFC).
"""
import os
import re
import json
import sqlite3
import hashlib
import secrets
from datetime import date
from functools import wraps
from io import BytesIO
from typing import NamedTuple
from werkzeug.utils import secure_filename
from flask import Flask, request, redirect, url_for, session, render_template, jsonify, send_from_directory, send_file

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024  # 5 MB uploads
DB_PATH = os.path.join(os.path.dirname(__file__), "bytecredits.db")
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "static", "uploads", "profiles")
UPLOAD_DIR_PAYMENTS = os.path.join(os.path.dirname(__file__), "static", "uploads", "payments")
CLASS_EXCEL_DIR = os.path.join(os.path.dirname(__file__), "static", "uploads", "class_rosters")
ATTENDANCE_EXCEL_DIR = os.path.join(os.path.dirname(__file__), "static", "uploads", "attendance_exports")
ALLOWED_PHOTO_EXT = {"png", "jpg", "jpeg", "gif", "webp"}

SUBJECT_LABELS = {
    "wt": "Web Technologies",
    "cd": "Compiler Design",
    "cn": "Computer Networks",
    "se": "Software Engineering",
}

CLASS_LABELS = {
    "cse-a": "CSE - A (3rd Year)",
    "cse-b": "CSE - B (3rd Year)",
    "ece-a": "ECE - A (2nd Year)",
}

ROLE_STUDENT = "student"
ROLE_TEACHER = "teacher"
ROLE_ADMIN = "admin"
_VALID_ROLES = frozenset({ROLE_STUDENT, ROLE_TEACHER, ROLE_ADMIN})

# Used only when BYTECREDITS_ADMIN_PASSWORD is unset (local/demo). Override in production.
SAMPLE_ADMIN_PASSWORD = "bytecredits-admin"


def get_admin_password() -> str:
    env = os.environ.get("BYTECREDITS_ADMIN_PASSWORD", "").strip()
    return env if env else SAMPLE_ADMIN_PASSWORD


def admin_using_sample_password() -> bool:
    return not bool(os.environ.get("BYTECREDITS_ADMIN_PASSWORD", "").strip())


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _migrate_attendance_semester_schema(conn: sqlite3.Connection) -> None:
    """Add semester_no to classes; rebuild subjects table if needed (drop UNIQUE on subject_name)."""
    cls_cols = {r[1] for r in conn.execute("PRAGMA table_info(attendance_classes)").fetchall()}
    if "semester_no" not in cls_cols:
        try:
            conn.execute("ALTER TABLE attendance_classes ADD COLUMN semester_no INTEGER NOT NULL DEFAULT 1")
            conn.commit()
        except sqlite3.OperationalError:
            pass
    if "academic_year" not in cls_cols:
        try:
            conn.execute("ALTER TABLE attendance_classes ADD COLUMN academic_year TEXT NOT NULL DEFAULT ''")
            conn.commit()
        except sqlite3.OperationalError:
            pass
    sub_cols_raw = conn.execute("PRAGMA table_info(attendance_subjects)").fetchall()
    sub_names = {r[1] for r in sub_cols_raw}
    if "semester_no" in sub_names:
        return
    has_created_at = "created_at" in sub_names
    conn.execute("BEGIN IMMEDIATE")
    try:
        conn.execute(
            """CREATE TABLE attendance_subjects_new (
                subject_key TEXT PRIMARY KEY,
                subject_name TEXT NOT NULL,
                semester_no INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )"""
        )
        if has_created_at:
            conn.execute(
                """INSERT INTO attendance_subjects_new (subject_key, subject_name, semester_no, created_at)
                   SELECT subject_key, subject_name, 1, COALESCE(created_at, CURRENT_TIMESTAMP)
                   FROM attendance_subjects"""
            )
        else:
            conn.execute(
                """INSERT INTO attendance_subjects_new (subject_key, subject_name, semester_no)
                   SELECT subject_key, subject_name, 1 FROM attendance_subjects"""
            )
        conn.execute("DROP TABLE attendance_subjects")
        conn.execute("ALTER TABLE attendance_subjects_new RENAME TO attendance_subjects")
        conn.commit()
    except sqlite3.Error:
        conn.rollback()
        raise


def init_db():
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rollno TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                department TEXT NOT NULL,
                year INTEGER NOT NULL,
                phone TEXT NOT NULL,
                email TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'student',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS cards (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rollno TEXT NOT NULL UNIQUE,
                card_uid TEXT NOT NULL UNIQUE,
                card_pin_hash TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (rollno) REFERENCES users(rollno)
            );
        """)
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS activities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rollno TEXT NOT NULL,
                description TEXT NOT NULL,
                amount INTEGER NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS profiles (
                rollno TEXT PRIMARY KEY,
                tagline TEXT,
                summary TEXT,
                skills TEXT,
                skills_json TEXT NOT NULL DEFAULT '{}',
                non_tech_skills TEXT,
                projects TEXT,
                achievements_json TEXT NOT NULL DEFAULT '[]',
                work_experience_json TEXT NOT NULL DEFAULT '[]',
                education_json TEXT NOT NULL DEFAULT '[]',
                location TEXT,
                linkedin TEXT,
                github TEXT,
                portfolio TEXT,
                is_public INTEGER NOT NULL DEFAULT 0,
                photo_url TEXT,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS credit_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rollno TEXT NOT NULL,
                amount INTEGER NOT NULL,
                screenshot_path TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                reviewed_at TEXT,
                review_note TEXT,
                FOREIGN KEY (rollno) REFERENCES users(rollno)
            );
            CREATE TABLE IF NOT EXISTS attendance_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                teacher_rollno TEXT NOT NULL,
                class_key TEXT,
                period TEXT,
                subject_key TEXT,
                student_rollno TEXT NOT NULL,
                card_uid TEXT,
                session_date TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS attendance_classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name_key TEXT UNIQUE NOT NULL,
                display_name TEXT NOT NULL,
                academic_year TEXT NOT NULL DEFAULT '',
                year INTEGER NOT NULL,
                semester_no INTEGER NOT NULL DEFAULT 1,
                dept_section TEXT NOT NULL,
                roll_prefix TEXT NOT NULL,
                start_seq INTEGER NOT NULL,
                end_seq INTEGER NOT NULL,
                pad_width INTEGER NOT NULL DEFAULT 2,
                missing_json TEXT NOT NULL DEFAULT '[]',
                excel_path TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS attendance_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                teacher_rollno TEXT NOT NULL,
                class_key TEXT NOT NULL,
                period TEXT NOT NULL,
                session_date TEXT NOT NULL,
                subject_key TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(class_key, period, session_date)
            );
            CREATE TABLE IF NOT EXISTS attendance_subjects (
                subject_key TEXT PRIMARY KEY,
                subject_name TEXT NOT NULL,
                semester_no INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS teacher_class_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                teacher_rollno TEXT NOT NULL,
                class_key TEXT NOT NULL,
                is_incharge INTEGER NOT NULL DEFAULT 0,
                UNIQUE(teacher_rollno, class_key)
            );
            CREATE TABLE IF NOT EXISTS teacher_subject_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                teacher_rollno TEXT NOT NULL,
                subject_key TEXT NOT NULL,
                UNIQUE(teacher_rollno, subject_key)
            );
        """)
        for col, typ in [("wallet_balance", "INTEGER DEFAULT 0"), ("attendance_pct", "REAL DEFAULT 0")]:
            try:
                conn.execute(f"ALTER TABLE users ADD COLUMN {col} {typ}")
                conn.commit()
            except sqlite3.OperationalError:
                pass
        try:
            conn.execute("ALTER TABLE cards ADD COLUMN card_pin_hash TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE profiles ADD COLUMN non_tech_skills TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE profiles ADD COLUMN skills_json TEXT NOT NULL DEFAULT '{}'")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE profiles ADD COLUMN achievements_json TEXT NOT NULL DEFAULT '[]'")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE profiles ADD COLUMN work_experience_json TEXT NOT NULL DEFAULT '[]'")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE profiles ADD COLUMN education_json TEXT NOT NULL DEFAULT '[]'")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE profiles ADD COLUMN is_public INTEGER NOT NULL DEFAULT 0")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'student'")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE users ADD COLUMN class_incharge_key TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        conn.execute("UPDATE users SET role = 'student' WHERE role IS NULL OR trim(role) = ''")
        _migrate_attendance_semester_schema(conn)
        for sk, sn in SUBJECT_LABELS.items():
            conn.execute(
                "INSERT OR IGNORE INTO attendance_subjects (subject_key, subject_name, semester_no) VALUES (?, ?, 1)",
                (sk, sn),
            )
        old_rows = conn.execute(
            """SELECT rollno, class_incharge_key
               FROM users
               WHERE lower(COALESCE(role, 'student')) = ?
                 AND class_incharge_key IS NOT NULL
                 AND trim(class_incharge_key) != ''""",
            (ROLE_TEACHER,),
        ).fetchall()
        for rw in old_rows:
            conn.execute(
                """INSERT OR IGNORE INTO teacher_class_assignments (teacher_rollno, class_key, is_incharge)
                   VALUES (?, ?, 1)""",
                ((rw["rollno"] or "").strip().upper(), (rw["class_incharge_key"] or "").strip()),
            )
        conn.commit()
        try:
            conn.execute("ALTER TABLE attendance_records ADD COLUMN session_date TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(UPLOAD_DIR_PAYMENTS, exist_ok=True)
    os.makedirs(CLASS_EXCEL_DIR, exist_ok=True)
    os.makedirs(ATTENDANCE_EXCEL_DIR, exist_ok=True)


def hash_password(pwd: str) -> str:
    return hashlib.sha256(pwd.strip().encode()).hexdigest()


def normalize_card_uid(uid: str) -> str:
    """Alphanumeric uppercase — matches UIDs whether stored with colons/dashes or not."""
    return "".join(c for c in (uid or "").strip().upper() if c.isalnum())


def make_class_name_key(year: int, dept_section: str) -> str:
    ds = (dept_section or "").strip().lower().replace(" ", "_")
    ds = re.sub(r"[^a-z0-9\-_]", "", ds)
    if not ds:
        raise ValueError("Department / section is invalid")
    return f"{int(year)}_{ds}"


def slug_class_name_label(class_name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (class_name or "").strip().lower()).strip("-")
    return s or "class"


def make_class_stem(year: int, semester_no: int, class_name: str) -> str:
    sem = int(semester_no)
    if not 1 <= sem <= 8:
        raise ValueError("Semester must be between 1 and 8")
    slug = slug_class_name_label(class_name)
    return f"{int(year)}_{sem}_{slug}"


def make_attendance_class_key(academic_year: str, class_name: str) -> str:
    """Key for attendance_classes.name_key: <academic_year>_<class-slug>.

    Example: 2025-26_cse-a
    """
    ay = (academic_year or "").strip().lower().replace(" ", "-")
    ay = re.sub(r"[^a-z0-9\-_]", "", ay)
    if not ay:
        raise ValueError("Academic year is invalid")
    slug = slug_class_name_label(class_name)
    return f"{ay}_{slug}"


def _parse_missing_piece(piece: str, out: set[int]) -> None:
    """Single token: '12' or '5-10' (no spaces inside the range)."""
    p = (piece or "").strip().replace("–", "-").replace("—", "-")
    if not p:
        return
    m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", p)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        lo, hi = (a, b) if a <= b else (b, a)
        out.update(range(lo, hi + 1))
        return
    try:
        out.add(int(p))
    except ValueError:
        pass


def _parse_missing_chunk(chunk: str, out: set[int]) -> None:
    """One comma/line segment: e.g. '5-10', '5 - 10', '3', '3 7 12', or '3, 5-8'."""
    c = (chunk or "").strip()
    if not c:
        return
    c = c.replace("–", "-").replace("—", "-")
    # Whole segment is one range with spaces around dash
    m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", c)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        lo, hi = (a, b) if a <= b else (b, a)
        out.update(range(lo, hi + 1))
        return
    for part in c.split():
        _parse_missing_piece(part, out)


def parse_missing_numbers(raw) -> set[int]:
    if raw is None:
        return set()
    if isinstance(raw, list):
        merged: set[int] = set()
        for x in raw:
            if x is None:
                continue
            merged |= parse_missing_numbers(str(x).strip())
        return merged
    s = str(raw).strip()
    if not s:
        return set()
    out: set[int] = set()
    # Split on newlines, commas, semicolons (not on hyphen — so "5-10" stays one chunk)
    for chunk in re.split(r"[\n,;]+", s):
        _parse_missing_chunk(chunk, out)
    return out


def build_roll_numbers(prefix: str, start_seq: int, end_seq: int, missing: set[int], pad_width: int | None) -> list[str]:
    p = (prefix or "").strip()
    if start_seq > end_seq:
        raise ValueError("Starting number must be ≤ ending number")
    if pad_width is None or pad_width < 1:
        pad_width = max(2, len(str(abs(start_seq))), len(str(abs(end_seq))))
    rolls = []
    for n in range(start_seq, end_seq + 1):
        if n in missing:
            continue
        rolls.append(f"{p}{n:0{pad_width}d}")
    return rolls


def _class_keys_for_student(conn: sqlite3.Connection, rollno: str) -> list[str]:
    """Find attendance class keys where this roll number belongs to roster range."""
    target = (rollno or "").strip().upper()
    if not target:
        return []
    rows = conn.execute(
        """SELECT name_key, roll_prefix, start_seq, end_seq, pad_width, missing_json
           FROM attendance_classes"""
    ).fetchall()
    out: list[str] = []
    for rw in rows:
        try:
            missing = parse_missing_numbers(rw["missing_json"] or "")
            rolls = build_roll_numbers(
                rw["roll_prefix"],
                int(rw["start_seq"]),
                int(rw["end_seq"]),
                missing,
                int(rw["pad_width"] or 0),
            )
            if target in {r.upper() for r in rolls}:
                out.append((rw["name_key"] or "").strip())
        except Exception:
            continue
    return [k for k in out if k]


def _teacher_incharge_class(conn: sqlite3.Connection, rollno: str) -> str:
    row = conn.execute(
        "SELECT COALESCE(NULLIF(trim(class_incharge_key), ''), '') AS class_incharge_key FROM users WHERE rollno = ?",
        (rollno,),
    ).fetchone()
    return (row["class_incharge_key"] if row else "") or ""


def _resolve_class_key(conn: sqlite3.Connection, key_or_label: str) -> str:
    """Map stored class key/label to attendance_classes.name_key when possible."""
    raw = (key_or_label or "").strip()
    if not raw:
        return ""
    # Already a real class key?
    rw = conn.execute(
        "SELECT name_key FROM attendance_classes WHERE name_key = ?",
        (raw,),
    ).fetchone()
    if rw and (rw["name_key"] or "").strip():
        return (rw["name_key"] or "").strip()
    # Sometimes older data stored dept_section/slug (e.g. 'cse-c')
    slug = slug_class_name_label(raw)
    rw = conn.execute(
        "SELECT name_key FROM attendance_classes WHERE dept_section = ? ORDER BY id DESC LIMIT 1",
        (slug,),
    ).fetchone()
    if rw and (rw["name_key"] or "").strip():
        return (rw["name_key"] or "").strip()
    return ""


def _teacher_assignments(conn: sqlite3.Connection, rollno: str) -> tuple[set[str], set[str], set[str]]:
    tr = (rollno or "").strip().upper()
    if not tr:
        return set(), set(), set()
    class_rows = conn.execute(
        """SELECT class_key, COALESCE(is_incharge, 0) AS is_incharge
           FROM teacher_class_assignments
           WHERE teacher_rollno = ?""",
        (tr,),
    ).fetchall()
    sub_rows = conn.execute(
        "SELECT subject_key FROM teacher_subject_assignments WHERE teacher_rollno = ?",
        (tr,),
    ).fetchall()
    assigned_classes = {(r["class_key"] or "").strip() for r in class_rows if (r["class_key"] or "").strip()}
    incharge_classes = {
        (r["class_key"] or "").strip()
        for r in class_rows
        if (r["class_key"] or "").strip() and int(r["is_incharge"] or 0) == 1
    }
    # Backward-compatible: also honor users.class_incharge_key (older data / partial updates).
    ck = _resolve_class_key(conn, (_teacher_incharge_class(conn, tr) or "").strip())
    if ck:
        incharge_classes.add(ck)
        assigned_classes.add(ck)
    assigned_subjects = {(r["subject_key"] or "").strip() for r in sub_rows if (r["subject_key"] or "").strip()}
    return incharge_classes, assigned_classes, assigned_subjects


def _attendance_subject_options(conn: sqlite3.Connection, semester_no: int | None = None) -> list[dict]:
    if semester_no is not None:
        rows = conn.execute(
            """SELECT subject_key, subject_name, COALESCE(semester_no, 1) AS semester_no
               FROM attendance_subjects WHERE semester_no = ?
               ORDER BY subject_name COLLATE NOCASE""",
            (int(semester_no),),
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT subject_key, subject_name, COALESCE(semester_no, 1) AS semester_no
               FROM attendance_subjects
               ORDER BY semester_no, subject_name COLLATE NOCASE"""
        ).fetchall()
    return [
        {"key": r["subject_key"], "name": r["subject_name"], "semester_no": int(r["semester_no"] or 1)}
        for r in rows
    ]


def _class_semester_no(conn: sqlite3.Connection, class_key: str) -> int | None:
    if not class_key:
        return None
    row = conn.execute(
        "SELECT COALESCE(semester_no, 1) AS semester_no FROM attendance_classes WHERE name_key = ?",
        (class_key,),
    ).fetchone()
    return int(row["semester_no"]) if row else None


def write_class_roster_excel(
    abs_path: str,
    roll_numbers: list[str],
    roll_to_name: dict[str, str] | None = None,
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "class_roster"
    ws.append(["Roll number", "Name"])
    for r in roll_numbers:
        nm = (roll_to_name or {}).get(r)
        nm = (nm or "").strip() or "-"
        ws.append([r, nm])
    wb.save(abs_path)


def _safe_excel_sheet_title(name: str, max_len: int = 31) -> str:
    """Excel worksheet names: max 31 chars; no [] : * ? / \\"""
    out = []
    for c in name:
        if c in "[]:*?/\\" or ord(c) < 32:
            out.append(" ")
        else:
            out.append(c)
    s = " ".join("".join(out).split())
    s = (s[:max_len] or "Sheet1").strip()
    return s or "Sheet1"


def _attendance_period_column(period: str) -> int | None:
    """Map period 1–9 to 1-based Excel column: P1→3, P2→4, … P9→11 (after roll + name)."""
    p = str(period or "").strip()
    if re.fullmatch(r"\d+", p):
        n = int(p)
    else:
        m = re.search(r"\d+", p)
        n = int(m.group(0)) if m else None
    if n is None or not (1 <= n <= 9):
        return None
    return 2 + n


def _attendance_excel_periods(class_key: str, session_date: str) -> list[dict]:
    """Read period status from class Excel date sheet (P1..P9 => columns 3..11)."""
    if not class_key or not session_date:
        return []
    with get_db() as conn:
        cls = conn.execute(
            "SELECT excel_path FROM attendance_classes WHERE name_key = ?",
            (class_key,),
        ).fetchone()
    if not cls or not (cls["excel_path"] or "").strip():
        return []
    rel = (cls["excel_path"] or "").replace("\\", "/").strip("/")
    if not rel or ".." in rel:
        return []
    abs_path = os.path.join(os.path.dirname(__file__), "static", rel.replace("/", os.sep))
    if not os.path.isfile(abs_path):
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(abs_path)
    except Exception:
        return []
    sheet_name = _safe_excel_sheet_title(session_date[:10])
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    out: list[dict] = []
    for p in range(1, 10):
        col = 2 + p
        header = ""
        if ws is not None:
            header = str(ws.cell(row=1, column=col).value or "").strip()
        out.append(
            {
                "period": str(p),
                "label": header or f"{p}",
                "taken": bool(header),
            }
        )
    return out


def write_attendance_session_excel(
    class_key: str,
    period: str,
    session_date: str,
    subject_key: str | None,
    rows: list[tuple[str, str]],
) -> str | None:
    """Write attendance into the class workbook created during class creation."""
    from openpyxl import load_workbook

    if not rows:
        return None
    with get_db() as conn:
        cls = conn.execute(
            "SELECT excel_path FROM attendance_classes WHERE name_key = ?",
            (class_key,),
        ).fetchone()
    if not cls or not (cls["excel_path"] or "").strip():
        return None

    rel = (cls["excel_path"] or "").replace("\\", "/").strip("/")
    if not rel or ".." in rel:
        return None
    abs_path = os.path.join(os.path.dirname(__file__), "static", rel.replace("/", os.sep))
    if not os.path.isfile(abs_path):
        return None

    wb = load_workbook(abs_path)
    base_ws = wb["class_roster"] if "class_roster" in wb.sheetnames else wb[wb.sheetnames[0]]
    roster: list[tuple[str, str]] = []
    for r in range(2, base_ws.max_row + 1):
        roll = str(base_ws.cell(row=r, column=1).value or "").strip().upper()
        if not roll:
            continue
        name = str(base_ws.cell(row=r, column=2).value or "").strip()
        roster.append((roll, name))

    date_part = session_date[:10] if re.match(r"^\d{4}-\d{2}-\d{2}", session_date) else secure_filename(session_date)[:32]
    sheet_name = _safe_excel_sheet_title(date_part)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value="Roll number")
        ws.cell(row=1, column=2, value="Student name")
        for idx, (roll, name) in enumerate(roster, start=2):
            ws.cell(row=idx, column=1, value=roll)
            ws.cell(row=idx, column=2, value=name)

    subject_name = SUBJECT_LABELS.get((subject_key or "").strip(), (subject_key or "").strip() or "Attendance")
    period_label = str(period or "").strip()
    subject_col_header = f"{subject_name} (P{period_label})" if period_label else subject_name
    subject_col = _attendance_period_column(period)
    if subject_col is None:
        return None
    ws.cell(row=1, column=subject_col, value=subject_col_header)

    present_rolls = {roll.strip().upper() for roll, _ in rows if (roll or "").strip()}
    roster_rows: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        roll = str(ws.cell(row=r, column=1).value or "").strip().upper()
        if roll:
            roster_rows[roll] = r
    for roll, _ in roster:
        rr = roster_rows.get(roll)
        if rr is not None:
            ws.cell(row=rr, column=subject_col, value="P" if roll in present_rolls else "A")

    wb.save(abs_path)
    return None


def get_user_by_rollno(rollno: str):
    with get_db() as conn:
        return conn.execute(
            "SELECT name, rollno, department, year, phone, email, COALESCE(wallet_balance, 0) AS wallet_balance FROM users WHERE rollno = ?",
            (rollno,),
        ).fetchone()


def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapped


def session_role() -> str:
    r = (session.get("role") or ROLE_STUDENT).lower()
    return r if r in _VALID_ROLES else ROLE_STUDENT


def redirect_home_for_session():
    r = session_role()
    if r == ROLE_ADMIN:
        return redirect(url_for("admin_dashboard"))
    if r == ROLE_TEACHER:
        return redirect(url_for("teacher_dashboard"))
    return redirect(url_for("student_dashboard"))


def is_admin_session() -> bool:
    if session.get("admin_ok"):
        return True
    if session.get("user_id") and session_role() == ROLE_ADMIN:
        return True
    return False


def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not is_admin_session():
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return wrapped


def student_only_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        if session_role() != ROLE_STUDENT:
            return redirect_home_for_session()
        return f(*args, **kwargs)
    return wrapped


def teacher_only_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        if session_role() != ROLE_TEACHER:
            return redirect_home_for_session()
        return f(*args, **kwargs)
    return wrapped


def student_or_teacher_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        if session_role() == ROLE_ADMIN:
            return redirect(url_for("admin_dashboard"))
        return f(*args, **kwargs)
    return wrapped


def teacher_or_admin_required(f):
    """Take attendance: teachers and logged-in admins only."""

    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        role = session_role()
        if role not in (ROLE_TEACHER, ROLE_ADMIN):
            return redirect_home_for_session()
        if role == ROLE_TEACHER and not user_has_card(session.get("rollno") or ""):
            return redirect(url_for("register_card"))
        return f(*args, **kwargs)
    return wrapped


@app.route("/")
def index():
    if session.get("user_id"):
        return redirect_home_for_session()
    return redirect(url_for("landing"))


@app.route("/landing")
def landing():
    return render_template("index.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "GET":
        return render_template("register.html")
    data = request.get_json() or request.form
    rollno = (data.get("rollno") or "").strip().upper()
    name = (data.get("username") or data.get("name") or "").strip()
    department = (data.get("department") or "").strip()
    year = data.get("year")
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()
    if not all([rollno, name, department, year, phone, email, password]):
        return jsonify({"ok": False, "error": "All fields required"}), 400
    try:
        year = int(year)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid year"}), 400
    if len(password) < 6:
        return jsonify({"ok": False, "error": "Password must be at least 6 characters"}), 400
    pwd_hash = hash_password(password)
    with get_db() as conn:
        try:
            conn.execute(
                """INSERT INTO users (rollno, name, department, year, phone, email, password_hash, role)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (rollno, name, department, year, phone, email, pwd_hash, ROLE_STUDENT),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            return jsonify({"ok": False, "error": "Roll number already registered"}), 409
    return jsonify({"ok": True, "redirect": url_for("login")})


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")
    data = request.get_json() or request.form
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not username or not password:
        return jsonify({"ok": False, "error": "Username and password required"}), 400
    with get_db() as conn:
        row = conn.execute(
            """SELECT id, rollno, name, COALESCE(NULLIF(trim(role), ''), ?) AS role
               FROM users WHERE (rollno = ? OR name = ?) AND password_hash = ?""",
            (ROLE_STUDENT, username.upper(), username, hash_password(password)),
        ).fetchone()
    if not row:
        return jsonify({"ok": False, "error": "Invalid credentials"}), 401
    role = (row["role"] or ROLE_STUDENT).lower()
    if role not in _VALID_ROLES:
        role = ROLE_STUDENT
    session["user_id"] = row["id"]
    session["rollno"] = row["rollno"]
    session["name"] = row["name"]
    session["role"] = role

    if role == ROLE_ADMIN:
        return jsonify({"ok": True, "redirect": url_for("admin_dashboard"), "needs_card": False})
    if role == ROLE_TEACHER:
        return jsonify({
            "ok": True,
            "redirect": url_for("teacher_dashboard"),
            "needs_card": not user_has_card(row["rollno"]),
        })
    return jsonify({
        "ok": True,
        "redirect": url_for("student_dashboard"),
        "needs_card": not user_has_card(row["rollno"]),
    })


def user_has_card(rollno: str) -> bool:
    with get_db() as conn:
        return conn.execute("SELECT 1 FROM cards WHERE rollno = ?", (rollno,)).fetchone() is not None


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/student-dashboard")
@student_only_required
def student_dashboard():
    if not user_has_card(session["rollno"]):
        return redirect(url_for("register_card"))
    rollno = session["rollno"]
    with get_db() as conn:
        row = conn.execute(
            "SELECT name, rollno, department, year, COALESCE(wallet_balance, 0) AS wallet_balance, COALESCE(attendance_pct, 0) AS attendance_pct FROM users WHERE rollno = ?",
            (rollno,),
        ).fetchone()
        class_keys = _class_keys_for_student(conn, rollno)
        present_periods = 0
        total_taken_periods = 0
        if class_keys:
            placeholders = ",".join("?" for _ in class_keys)
            total_taken_periods = int(
                conn.execute(
                    f"""SELECT COUNT(*) FROM attendance_sessions
                        WHERE class_key IN ({placeholders})""",
                    class_keys,
                ).fetchone()[0]
                or 0
            )
            present_periods = int(
                conn.execute(
                    f"""SELECT COUNT(*) FROM attendance_records
                        WHERE student_rollno = ?
                          AND class_key IN ({placeholders})""",
                    [rollno] + class_keys,
                ).fetchone()[0]
                or 0
            )
        else:
            # Fallback: compute from whichever sessions include this student's records.
            present_periods = int(
                conn.execute(
                    "SELECT COUNT(*) FROM attendance_records WHERE student_rollno = ?",
                    (rollno,),
                ).fetchone()[0]
                or 0
            )
            total_taken_periods = int(
                conn.execute(
                    """SELECT COUNT(*)
                       FROM attendance_sessions s
                       WHERE EXISTS (
                           SELECT 1 FROM attendance_records ar
                           WHERE ar.student_rollno = ? AND ar.class_key = s.class_key
                       )""",
                    (rollno,),
                ).fetchone()[0]
                or 0
            )
        attendance_pct_live = (present_periods * 100.0 / total_taken_periods) if total_taken_periods > 0 else 0.0
        activities = conn.execute(
            "SELECT description, amount, created_at FROM activities WHERE rollno = ? ORDER BY created_at DESC LIMIT 10",
            (rollno,),
        ).fetchall()
    if not row:
        return redirect(url_for("logout"))
    acts = [{"description": r["description"], "amount": r["amount"], "created_at": r["created_at"]} for r in activities]
    return render_template(
        "student_dashboard.html",
        name=row["name"],
        rollno=row["rollno"],
        department=row["department"],
        year=row["year"],
        wallet_balance=int(row["wallet_balance"]) if row["wallet_balance"] is not None else 0,
        attendance_pct=attendance_pct_live,
        attendance_present_periods=present_periods,
        attendance_total_periods=total_taken_periods,
        activities=acts,
    )


@app.route("/teacher-dashboard")
@teacher_only_required
def teacher_dashboard():
    if not user_has_card(session["rollno"]):
        return redirect(url_for("register_card"))
    rollno = session["rollno"]
    with get_db() as conn:
        row = conn.execute(
            "SELECT name, rollno, department, year, COALESCE(wallet_balance, 0) AS wallet_balance, COALESCE(attendance_pct, 0) AS attendance_pct FROM users WHERE rollno = ?",
            (rollno,),
        ).fetchone()
        activities = conn.execute(
            "SELECT description, amount, created_at FROM activities WHERE rollno = ? ORDER BY created_at DESC LIMIT 10",
            (rollno,),
        ).fetchall()
    if not row:
        return redirect(url_for("logout"))
    acts = [{"description": r["description"], "amount": r["amount"], "created_at": r["created_at"]} for r in activities]
    return render_template(
        "teacher_dashboard.html",
        name=row["name"],
        rollno=row["rollno"],
        department=row["department"],
        year=row["year"],
        wallet_balance=int(row["wallet_balance"]) if row["wallet_balance"] is not None else 0,
        attendance_pct=float(row["attendance_pct"]) if row["attendance_pct"] is not None else 0,
        activities=acts,
    )


@app.route("/register-card", methods=["GET", "POST"])
@student_or_teacher_required
def register_card():
    rollno = session.get("rollno")
    if not rollno:
        return redirect(url_for("login"))
    if user_has_card(rollno):
        return redirect_home_for_session()
    if request.method == "GET":
        return render_template("register_card.html", rollno=rollno, name=session.get("name"))
    data = request.get_json() or request.form
    card_uid = (data.get("card_uid") or "").strip()
    card_pin = (data.get("card_pin") or "").strip()
    rollno_submit = (data.get("rollno") or "").strip().upper()
    if not card_uid:
        return jsonify({"ok": False, "error": "Card UID required (tap card via NFC)"}), 400
    if len(card_pin) != 4 or not card_pin.isdigit():
        return jsonify({"ok": False, "error": "Card PIN must be 4 digits"}), 400
    if rollno_submit != rollno:
        return jsonify({"ok": False, "error": "Roll number does not match logged-in user"}), 403
    card_pin_hash = hash_password(card_pin)
    with get_db() as conn:
        try:
            conn.execute(
                "INSERT INTO cards (rollno, card_uid, card_pin_hash) VALUES (?, ?, ?)",
                (rollno, card_uid, card_pin_hash),
            )
            conn.commit()
        except sqlite3.IntegrityError as e:
            if "rollno" in str(e).lower() or "UNIQUE" in str(e):
                return jsonify({"ok": False, "error": "This roll number already has a card"}), 409
            return jsonify({"ok": False, "error": "This card is already registered to another user"}), 409
    return jsonify({"ok": True, "redirect": url_for("teacher_dashboard" if session_role() == ROLE_TEACHER else "student_dashboard")})


@app.route("/api/check-card")
@student_or_teacher_required
def api_check_card():
    return jsonify({"has_card": user_has_card(session["rollno"])})


@app.route("/api/nfc-pay", methods=["POST"])
def api_nfc_pay():
    """Debit payer (card + PIN) and credit recipient roll number."""
    data = request.get_json() or {}
    recipient = (data.get("recipient_rollno") or "").strip().upper()
    card_uid = (data.get("card_uid") or "").strip()
    card_pin = (data.get("card_pin") or "").strip()
    try:
        amount = int(data.get("amount"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid amount"}), 400
    if amount < 1:
        return jsonify({"ok": False, "error": "Amount must be at least 1 PTS"}), 400
    if not recipient or not card_uid:
        return jsonify({"ok": False, "error": "Recipient roll number and card UID required"}), 400
    if len(card_pin) != 4 or not card_pin.isdigit():
        return jsonify({"ok": False, "error": "Card PIN must be 4 digits"}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("BEGIN IMMEDIATE")
        norm = normalize_card_uid(card_uid)
        card = conn.execute(
            """SELECT rollno, card_pin_hash FROM cards WHERE
               upper(replace(replace(replace(trim(card_uid), '-', ''), ':', ''), ' ', '')) = ?""",
            (norm,),
        ).fetchone()
        if not card:
            conn.rollback()
            return jsonify({"ok": False, "error": "Card not registered"}), 404
        payer_roll = card["rollno"]
        if not card["card_pin_hash"] or hash_password(card_pin) != card["card_pin_hash"]:
            conn.rollback()
            return jsonify({"ok": False, "error": "Invalid card PIN"}), 403
        if payer_roll.upper() == recipient:
            conn.rollback()
            return jsonify({"ok": False, "error": "Cannot pay yourself"}), 400
        payee = conn.execute("SELECT rollno, name FROM users WHERE rollno = ?", (recipient,)).fetchone()
        if not payee:
            conn.rollback()
            return jsonify({"ok": False, "error": "Recipient roll number not found"}), 404
        payer_row = conn.execute(
            "SELECT name, COALESCE(wallet_balance, 0) AS wallet_balance FROM users WHERE rollno = ?",
            (payer_roll,),
        ).fetchone()
        if not payer_row:
            conn.rollback()
            return jsonify({"ok": False, "error": "Payer account missing"}), 500
        bal = int(payer_row["wallet_balance"] or 0)
        if bal < amount:
            conn.rollback()
            return jsonify({"ok": False, "error": "Insufficient balance", "balance": bal}), 402

        conn.execute(
            "UPDATE users SET wallet_balance = COALESCE(wallet_balance, 0) - ? WHERE rollno = ?",
            (amount, payer_roll),
        )
        conn.execute(
            "UPDATE users SET wallet_balance = COALESCE(wallet_balance, 0) + ? WHERE rollno = ?",
            (amount, recipient),
        )
        conn.execute(
            "INSERT INTO activities (rollno, description, amount) VALUES (?, ?, ?)",
            (payer_roll, f"NFC payment to {recipient} ({payee['name']})", -amount),
        )
        conn.execute(
            "INSERT INTO activities (rollno, description, amount) VALUES (?, ?, ?)",
            (recipient, f"NFC received from {payer_roll} ({payer_row['name']})", amount),
        )
        conn.commit()
        return jsonify({
            "ok": True,
            "message": f"Paid {amount} PTS to {recipient}",
            "payer_rollno": payer_roll,
            "new_balance_estimate": bal - amount,
        })
    except sqlite3.Error:
        conn.rollback()
        return jsonify({"ok": False, "error": "Database error"}), 500
    finally:
        conn.close()


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if is_admin_session():
        return redirect(url_for("admin_dashboard"))
    sample = admin_using_sample_password()
    if request.method == "POST":
        pwd = (request.form.get("password") or "").strip()
        if pwd and secrets.compare_digest(pwd, get_admin_password()):
            session["admin_ok"] = True
            return redirect(url_for("admin_dashboard"))
        return render_template(
            "admin_login.html",
            error="Invalid password",
            sample_mode=sample,
            sample_hint=SAMPLE_ADMIN_PASSWORD if sample else None,
        ), 401
    return render_template(
        "admin_login.html",
        error=None,
        sample_mode=sample,
        sample_hint=SAMPLE_ADMIN_PASSWORD if sample else None,
    )


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_ok", None)
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin_dashboard():
    with get_db() as conn:
        teachers = conn.execute(
            """SELECT id, rollno, name, department, email, phone, created_at,
                      COALESCE(role, 'student') AS role,
                      COALESCE(NULLIF(trim(class_incharge_key), ''), '') AS class_incharge_key
               FROM users WHERE lower(COALESCE(role, 'student')) IN (?, ?)
               ORDER BY role, name COLLATE NOCASE""",
            (ROLE_TEACHER, ROLE_ADMIN),
        ).fetchall()
        pending = conn.execute(
            """SELECT cr.id, cr.rollno, cr.amount, cr.screenshot_path, cr.created_at, u.name
               FROM credit_requests cr
               JOIN users u ON u.rollno = cr.rollno
               WHERE cr.status = 'pending'
               ORDER BY cr.created_at ASC"""
        ).fetchall()
        recent = conn.execute(
            """SELECT cr.id, cr.rollno, cr.amount, cr.status, cr.created_at, cr.reviewed_at, u.name
               FROM credit_requests cr
               JOIN users u ON u.rollno = cr.rollno
               WHERE cr.status != 'pending'
               ORDER BY cr.reviewed_at DESC, cr.id DESC
               LIMIT 30"""
        ).fetchall()
        ac = conn.execute(
            """SELECT id, name_key, display_name, COALESCE(NULLIF(trim(academic_year), ''), '') AS academic_year,
                      year, COALESCE(semester_no, 1) AS semester_no, dept_section,
                      roll_prefix, start_seq, end_seq, excel_path, created_at
               FROM attendance_classes ORDER BY display_name COLLATE NOCASE"""
        ).fetchall()
        attendance_classes_json = [
            {
                "id": int(r["id"]),
                "name_key": (r["name_key"] or ""),
                "display_name": (r["display_name"] or ""),
                "academic_year": (r["academic_year"] or ""),
                "semester_no": int(r["semester_no"] or 1),
            }
            for r in ac
        ]
        subjects = _attendance_subject_options(conn)
        tca_rows = conn.execute(
            """SELECT teacher_rollno, class_key, COALESCE(is_incharge, 0) AS is_incharge
               FROM teacher_class_assignments"""
        ).fetchall()
        tsa_rows = conn.execute(
            "SELECT teacher_rollno, subject_key FROM teacher_subject_assignments"
        ).fetchall()
    teacher_class_map: dict[str, list[str]] = {}
    teacher_incharge_map: dict[str, str] = {}
    for rw in tca_rows:
        tr = (rw["teacher_rollno"] or "").upper()
        ck = (rw["class_key"] or "").strip()
        if not tr or not ck:
            continue
        teacher_class_map.setdefault(tr, []).append(ck)
        if int(rw["is_incharge"] or 0) == 1:
            teacher_incharge_map[tr] = ck
    teacher_subject_map: dict[str, list[str]] = {}
    for rw in tsa_rows:
        tr = (rw["teacher_rollno"] or "").upper()
        sk = (rw["subject_key"] or "").strip()
        if tr and sk:
            teacher_subject_map.setdefault(tr, []).append(sk)
    password_admin_only = bool(session.get("admin_ok")) and not bool(session.get("user_id"))
    return render_template(
        "admin_dashboard.html",
        pending=pending,
        recent=recent,
        teachers=teachers,
        attendance_classes=ac,
        attendance_classes_json=attendance_classes_json,
        attendance_subjects=subjects,
        teacher_class_map=teacher_class_map,
        teacher_incharge_map=teacher_incharge_map,
        teacher_subject_map=teacher_subject_map,
        password_admin_only=password_admin_only,
    )


@app.route("/admin/teacher", methods=["POST"])
@admin_required
def admin_add_teacher():
    data = request.get_json(silent=True) or request.form
    staff_id = (data.get("staff_id") or data.get("rollno") or "").strip().upper()
    name = (data.get("name") or "").strip()
    department = (data.get("department") or "").strip()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()
    role_raw = (data.get("role") or ROLE_TEACHER).strip().lower()
    new_role = ROLE_ADMIN if role_raw == ROLE_ADMIN else ROLE_TEACHER
    class_incharge_key = (data.get("class_incharge_key") or "").strip()
    assigned_class_keys = data.get("assigned_class_keys") or []
    assigned_subject_keys = data.get("assigned_subject_keys") or []
    if isinstance(assigned_class_keys, str):
        assigned_class_keys = [x.strip() for x in assigned_class_keys.split(",") if x.strip()]
    if isinstance(assigned_subject_keys, str):
        assigned_subject_keys = [x.strip() for x in assigned_subject_keys.split(",") if x.strip()]
    if not isinstance(assigned_class_keys, list):
        assigned_class_keys = []
    if not isinstance(assigned_subject_keys, list):
        assigned_subject_keys = []
    assigned_class_keys = [str(x).strip() for x in assigned_class_keys if str(x).strip()]
    assigned_subject_keys = [str(x).strip() for x in assigned_subject_keys if str(x).strip()]
    if not all([staff_id, name, department, phone, email, password]):
        return jsonify({"ok": False, "error": "All fields are required"}), 400
    if len(password) < 6:
        return jsonify({"ok": False, "error": "Password must be at least 6 characters"}), 400
    if new_role != ROLE_TEACHER:
        class_incharge_key = ""
        assigned_class_keys = []
        assigned_subject_keys = []
    if class_incharge_key and class_incharge_key not in assigned_class_keys:
        assigned_class_keys.append(class_incharge_key)
    pwd_hash = hash_password(password)
    with get_db() as conn:
        if class_incharge_key:
            exists = conn.execute(
                "SELECT 1 FROM attendance_classes WHERE name_key = ?",
                (class_incharge_key,),
            ).fetchone()
            if not exists:
                return jsonify({"ok": False, "error": "Selected class does not exist"}), 400
        for ck in assigned_class_keys:
            exists = conn.execute("SELECT 1 FROM attendance_classes WHERE name_key = ?", (ck,)).fetchone()
            if not exists:
                return jsonify({"ok": False, "error": f"Assigned class not found: {ck}"}), 400
        for sk in assigned_subject_keys:
            exists = conn.execute("SELECT 1 FROM attendance_subjects WHERE subject_key = ?", (sk,)).fetchone()
            if not exists:
                return jsonify({"ok": False, "error": f"Assigned subject not found: {sk}"}), 400
        try:
            conn.execute(
                """INSERT INTO users (rollno, name, department, year, phone, email, password_hash, role, class_incharge_key)
                   VALUES (?, ?, ?, 0, ?, ?, ?, ?, ?)""",
                (staff_id, name, department, phone, email, pwd_hash, new_role, class_incharge_key or None),
            )
            for ck in assigned_class_keys:
                conn.execute(
                    """INSERT OR REPLACE INTO teacher_class_assignments (teacher_rollno, class_key, is_incharge)
                       VALUES (?, ?, ?)""",
                    (staff_id, ck, 1 if ck == class_incharge_key else 0),
                )
            for sk in assigned_subject_keys:
                conn.execute(
                    "INSERT OR IGNORE INTO teacher_subject_assignments (teacher_rollno, subject_key) VALUES (?, ?)",
                    (staff_id, sk),
                )
            conn.commit()
        except sqlite3.IntegrityError:
            return jsonify({"ok": False, "error": "Staff ID / roll number already in use"}), 409
    return jsonify({"ok": True, "redirect": url_for("admin_dashboard")})


@app.route("/admin/subject", methods=["POST"])
@admin_required
def admin_add_subject():
    data = request.get_json(silent=True) or request.form
    subject_name = (data.get("subject_name") or "").strip()
    subject_key = (data.get("subject_key") or "").strip().lower()
    try:
        semester_no = int(data.get("semester_no") or data.get("sem") or 0)
    except (TypeError, ValueError):
        semester_no = 0
    if not subject_name:
        return jsonify({"ok": False, "error": "Subject name is required"}), 400
    if not subject_key:
        subject_key = re.sub(r"[^a-z0-9]+", "-", subject_name.lower()).strip("-")
    if not subject_key:
        return jsonify({"ok": False, "error": "Invalid subject key"}), 400
    if not 1 <= semester_no <= 8:
        return jsonify({"ok": False, "error": "Semester must be between 1 and 8"}), 400
    with get_db() as conn:
        try:
            conn.execute(
                "INSERT INTO attendance_subjects (subject_key, subject_name, semester_no) VALUES (?, ?, ?)",
                (subject_key, subject_name, semester_no),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            return jsonify({"ok": False, "error": "Subject key already exists"}), 409
    return jsonify({"ok": True, "subject_key": subject_key})


@app.route("/admin/teacher/<int:user_id>/update", methods=["POST"])
@admin_required
def admin_update_teacher(user_id: int):
    data = request.get_json(silent=True) or request.form
    name = (data.get("name") or "").strip()
    department = (data.get("department") or "").strip()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    role_raw = (data.get("role") or ROLE_TEACHER).strip().lower()
    new_role = ROLE_ADMIN if role_raw == ROLE_ADMIN else ROLE_TEACHER
    class_incharge_key = (data.get("class_incharge_key") or "").strip()
    assigned_class_keys = data.get("assigned_class_keys") or []
    assigned_subject_keys = data.get("assigned_subject_keys") or []
    if isinstance(assigned_class_keys, str):
        assigned_class_keys = [x.strip() for x in assigned_class_keys.split(",") if x.strip()]
    if isinstance(assigned_subject_keys, str):
        assigned_subject_keys = [x.strip() for x in assigned_subject_keys.split(",") if x.strip()]
    if not isinstance(assigned_class_keys, list):
        assigned_class_keys = []
    if not isinstance(assigned_subject_keys, list):
        assigned_subject_keys = []
    assigned_class_keys = [str(x).strip() for x in assigned_class_keys if str(x).strip()]
    assigned_subject_keys = [str(x).strip() for x in assigned_subject_keys if str(x).strip()]
    if not all([name, department, phone, email]):
        return jsonify({"ok": False, "error": "Name, department, phone, and email are required"}), 400
    if new_role != ROLE_TEACHER:
        class_incharge_key = ""
        assigned_class_keys = []
        assigned_subject_keys = []
    if class_incharge_key and class_incharge_key not in assigned_class_keys:
        assigned_class_keys.append(class_incharge_key)
    with get_db() as conn:
        staff = conn.execute(
            """SELECT id, rollno, COALESCE(role, 'student') AS role
               FROM users WHERE id = ?""",
            (user_id,),
        ).fetchone()
        if not staff:
            return jsonify({"ok": False, "error": "Staff account not found"}), 404
        role_now = (staff["role"] or ROLE_STUDENT).strip().lower()
        if role_now not in (ROLE_TEACHER, ROLE_ADMIN):
            return jsonify({"ok": False, "error": "Only teacher/admin rows are editable here"}), 400
        if class_incharge_key:
            cls = conn.execute(
                "SELECT 1 FROM attendance_classes WHERE name_key = ?",
                (class_incharge_key,),
            ).fetchone()
            if not cls:
                return jsonify({"ok": False, "error": "Selected class does not exist"}), 400
        for ck in assigned_class_keys:
            cls = conn.execute("SELECT 1 FROM attendance_classes WHERE name_key = ?", (ck,)).fetchone()
            if not cls:
                return jsonify({"ok": False, "error": f"Assigned class not found: {ck}"}), 400
        for sk in assigned_subject_keys:
            sub = conn.execute("SELECT 1 FROM attendance_subjects WHERE subject_key = ?", (sk,)).fetchone()
            if not sub:
                return jsonify({"ok": False, "error": f"Assigned subject not found: {sk}"}), 400
        staff_roll = (staff["rollno"] or "").strip().upper()
        conn.execute(
            """UPDATE users
               SET name = ?, department = ?, phone = ?, email = ?, role = ?, class_incharge_key = ?
               WHERE id = ?""",
            (name, department, phone, email, new_role, class_incharge_key or None, user_id),
        )
        conn.execute("DELETE FROM teacher_class_assignments WHERE teacher_rollno = ?", (staff_roll,))
        conn.execute("DELETE FROM teacher_subject_assignments WHERE teacher_rollno = ?", (staff_roll,))
        for ck in assigned_class_keys:
            conn.execute(
                """INSERT OR REPLACE INTO teacher_class_assignments (teacher_rollno, class_key, is_incharge)
                   VALUES (?, ?, ?)""",
                (staff_roll, ck, 1 if ck == class_incharge_key else 0),
            )
        for sk in assigned_subject_keys:
            conn.execute(
                "INSERT OR IGNORE INTO teacher_subject_assignments (teacher_rollno, subject_key) VALUES (?, ?)",
                (staff_roll, sk),
            )
        conn.commit()
    return jsonify({"ok": True, "message": "Staff row updated"})


@app.route("/admin/class", methods=["POST"])
@admin_required
def admin_create_class():
    data = request.get_json(silent=True) or request.form
    academic_year = (data.get("academic_year") or data.get("acad_year") or data.get("ay") or "").strip()
    try:
        year = int(data.get("year") or 1)
        semester_no = int(data.get("semester_no") or data.get("sem") or 1)
        start_seq = int(data.get("start_seq") or data.get("start_pin"))
        end_seq = int(data.get("end_seq") or data.get("end_pin"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Start and end must be integers"}), 400
    class_name = (data.get("class_name") or data.get("dept_section") or "").strip()
    roll_prefix = (data.get("roll_prefix") or "").strip()
    display_name = (data.get("display_name") or "").strip()
    missing_raw = data.get("missing_numbers") or data.get("missing") or ""
    if not class_name or not roll_prefix:
        return jsonify({"ok": False, "error": "Class name and roll prefix are required"}), 400
    # Backward compatible: if academic_year isn't provided, fall back to numeric year.
    if not academic_year:
        academic_year = str(year)
    if not 1 <= semester_no <= 8:
        return jsonify({"ok": False, "error": "Semester must be between 1 and 8"}), 400
    if start_seq > end_seq:
        return jsonify({"ok": False, "error": "Start must be ≤ end"}), 400
    try:
        name_key = make_attendance_class_key(academic_year, class_name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    dept_section = slug_class_name_label(class_name)
    missing = parse_missing_numbers(missing_raw)
    if not display_name:
        display_name = f"{academic_year} — {class_name}"
    missing_json = json.dumps(sorted(missing))
    pw = max(2, len(str(start_seq)), len(str(end_seq)))
    with get_db() as conn:
        try:
            conn.execute(
                """INSERT INTO attendance_classes
                   (name_key, display_name, academic_year, year, semester_no, dept_section, roll_prefix, start_seq, end_seq, pad_width, missing_json, excel_path)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    name_key,
                    display_name,
                    academic_year,
                    year,
                    semester_no,
                    dept_section,
                    roll_prefix,
                    start_seq,
                    end_seq,
                    pw,
                    missing_json,
                    "",
                ),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            return jsonify({"ok": False, "error": "A class with this academic year and name already exists"}), 409
    return jsonify({"ok": True, "redirect": url_for("admin_dashboard"), "name_key": name_key})


@app.route("/admin/class/<int:class_id>/delete", methods=["POST"])
@admin_required
def admin_delete_class(class_id: int):
    with get_db() as conn:
        row = conn.execute(
            "SELECT excel_path FROM attendance_classes WHERE id = ?",
            (class_id,),
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Class not found"}), 404
        rel = row["excel_path"] or ""
        conn.execute("DELETE FROM attendance_classes WHERE id = ?", (class_id,))
        conn.commit()
    if rel:
        fp = os.path.join(os.path.dirname(__file__), "static", rel.replace("/", os.sep))
        try:
            if os.path.isfile(fp):
                os.remove(fp)
        except OSError:
            pass
    return jsonify({"ok": True})


@app.route("/admin/class/<int:class_id>/download")
@admin_required
def admin_download_class_roster(class_id: int):
    static_root = os.path.join(os.path.dirname(__file__), "static")
    with get_db() as conn:
        row = conn.execute(
            "SELECT excel_path, name_key, dept_section, academic_year, COALESCE(semester_no, 1) AS semester_no FROM attendance_classes WHERE id = ?",
            (class_id,),
        ).fetchone()
    if not row or not row["excel_path"]:
        return redirect(url_for("admin_dashboard"))
    rel = (row["excel_path"] or "").replace("\\", "/").strip("/")
    if not rel or ".." in rel:
        return redirect(url_for("admin_dashboard"))
    directory = os.path.dirname(rel)
    fname = os.path.basename(rel)
    send_dir = os.path.join(static_root, directory.replace("/", os.sep))
    if not os.path.isfile(os.path.join(send_dir, fname)):
        return redirect(url_for("admin_dashboard"))
    return send_from_directory(
        send_dir,
        fname,
        as_attachment=True,
        download_name=f"{(row['academic_year'] or '').strip() or row['name_key']}_sem{int(row['semester_no'] or 1)}_{secure_filename(row['dept_section'] or '') or 'class'}.xlsx",
    )


@app.route("/admin/class/<int:class_id>/excel/create", methods=["POST"])
@admin_required
def admin_create_class_excel_from_db(class_id: int, semester_override: int | None = None):
    """(Re)generate the class roster Excel using stored DB values."""
    with get_db() as conn:
        row = conn.execute(
            """SELECT id, name_key, academic_year, COALESCE(semester_no, 1) AS semester_no,
                      dept_section, roll_prefix, start_seq, end_seq, pad_width, missing_json, excel_path
               FROM attendance_classes WHERE id = ?""",
            (class_id,),
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Class not found"}), 404

        name_key = (row["name_key"] or "").strip()
        academic_year = (row["academic_year"] or "").strip()
        semester_no = int(semester_override) if semester_override is not None else int(row["semester_no"] or 1)
        if not 1 <= int(semester_no) <= 8:
            return jsonify({"ok": False, "error": "Semester must be between 1 and 8"}), 400
        dept_section = (row["dept_section"] or "").strip()
        roll_prefix = (row["roll_prefix"] or "").strip()
        try:
            start_seq = int(row["start_seq"])
            end_seq = int(row["end_seq"])
            pad_width = int(row["pad_width"]) if row["pad_width"] is not None else None
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "Invalid class roll range data"}), 400
        try:
            missing = set(int(x) for x in (json.loads(row["missing_json"] or "[]") or []))
        except Exception:
            missing = set()

        try:
            rolls = build_roll_numbers(roll_prefix, start_seq, end_seq, missing, pad_width)
        except ValueError as e:
            return jsonify({"ok": False, "error": str(e)}), 400
        if not rolls:
            return jsonify({"ok": False, "error": "No roll numbers after applying range and missing list"}), 400

        # Fill names from users table where available; otherwise '-'
        roll_to_name: dict[str, str] = {}
        if rolls:
            ph = ",".join(["?"] * len(rolls))
            user_rows = conn.execute(
                f"SELECT rollno, name FROM users WHERE rollno IN ({ph})",
                tuple(rolls),
            ).fetchall()
            for ur in user_rows:
                rn = (ur["rollno"] or "").strip()
                nm = (ur["name"] or "").strip()
                if rn:
                    roll_to_name[rn] = nm

        rel = (row["excel_path"] or "").replace("\\", "/").strip()
        # Always generate using academic_year + semester in filename (avoid collisions with class_id suffix).
        safe_ay = secure_filename(academic_year) or secure_filename(name_key) or f"class_{class_id}"
        safe_class = secure_filename(dept_section) or secure_filename(name_key) or f"class_{class_id}"
        filename = f"{safe_ay}_sem{semester_no}_{safe_class}.xlsx"
        rel = os.path.join("uploads", "class_rosters", filename).replace("\\", "/")
        if semester_override is not None:
            conn.execute(
                "UPDATE attendance_classes SET excel_path = ?, semester_no = ? WHERE id = ?",
                (rel, int(semester_no), class_id),
            )
        else:
            conn.execute("UPDATE attendance_classes SET excel_path = ? WHERE id = ?", (rel, class_id))
        conn.commit()

    abs_xlsx = os.path.join(os.path.dirname(__file__), "static", rel.replace("/", os.sep))
    os.makedirs(os.path.dirname(abs_xlsx), exist_ok=True)
    static_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "static"))
    if not os.path.abspath(abs_xlsx).startswith(static_root):
        return jsonify({"ok": False, "error": "Invalid file path"}), 400
    try:
        write_class_roster_excel(abs_xlsx, rolls, roll_to_name=roll_to_name)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Excel failed: {e}"}), 500
    return jsonify({"ok": True})


@app.route("/admin/class/excel/create", methods=["POST"])
@admin_required
def admin_create_excel_choose_class():
    data = request.get_json(silent=True) or request.form
    try:
        class_id = int(data.get("class_id") or 0)
    except (TypeError, ValueError):
        class_id = 0
    if class_id <= 0:
        return jsonify({"ok": False, "error": "Select a class"}), 400
    try:
        semester_no = int(data.get("semester_no"))
    except (TypeError, ValueError):
        semester_no = 0
    if not 1 <= semester_no <= 8:
        return jsonify({"ok": False, "error": "Semester must be between 1 and 8"}), 400
    resp = admin_create_class_excel_from_db(class_id, semester_override=semester_no)
    # If resp is (json, code) tuple, unwrap
    if isinstance(resp, tuple):
        body, code = resp
        if code != 200:
            return resp
    return jsonify({"ok": True, "download": url_for("admin_download_class_roster", class_id=class_id)})


@app.route("/api/view-attendance-subject-options", methods=["GET"])
def api_view_attendance_subject_options():
    """Subject dropdown options for /view-attendance based on selected class and session role."""
    password_only_admin = bool(session.get("admin_ok")) and not bool(session.get("user_id"))
    if not session.get("user_id") and not password_only_admin:
        return jsonify({"ok": False, "error": "Not logged in"}), 401
    r = ROLE_ADMIN if password_only_admin else session_role()
    rollno = "" if password_only_admin else (session.get("rollno") or "")
    if r not in (ROLE_TEACHER, ROLE_ADMIN):
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    class_key = (request.args.get("class_key") or "").strip()
    with get_db() as conn:
        class_sem = _class_semester_no(conn, class_key) if class_key else None
        cand = _attendance_subject_options(conn, class_sem)
        if r == ROLE_ADMIN or password_only_admin:
            opts = [{"value": s["key"], "label": f"{s['name']} (Sem {s['semester_no']})"} for s in cand]
            return jsonify({"ok": True, "subjects": opts})
        # Teacher
        incharge_classes, assigned_classes, assigned_subjects = _teacher_assignments(conn, rollno)
        teacher_selected_incharge = bool(class_key and class_key in incharge_classes)
        if teacher_selected_incharge:
            opts = [{"value": s["key"], "label": f"{s['name']} (Sem {s['semester_no']})"} for s in cand]
        else:
            opts = [
                {"value": s["key"], "label": f"{s['name']} (Sem {s['semester_no']})"}
                for s in cand
                if s["key"] in assigned_subjects
            ]
        return jsonify({"ok": True, "subjects": opts})


@app.route("/admin/credit/<int:req_id>/approve", methods=["POST"])
@admin_required
def admin_credit_approve(req_id: int):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("BEGIN IMMEDIATE")
        row = conn.execute(
            "SELECT id, rollno, amount, status FROM credit_requests WHERE id = ?",
            (req_id,),
        ).fetchone()
        if not row or row["status"] != "pending":
            conn.rollback()
            return jsonify({"ok": False, "error": "Request not found or already processed"}), 404
        conn.execute(
            "UPDATE users SET wallet_balance = COALESCE(wallet_balance, 0) + ? WHERE rollno = ?",
            (row["amount"], row["rollno"]),
        )
        conn.execute(
            """UPDATE credit_requests
               SET status = 'approved', reviewed_at = CURRENT_TIMESTAMP, review_note = ?
               WHERE id = ?""",
            ((request.form.get("note") or "").strip() or None, req_id),
        )
        conn.execute(
            "INSERT INTO activities (rollno, description, amount) VALUES (?, ?, ?)",
            (row["rollno"], f"Credit top-up approved (request #{req_id})", row["amount"]),
        )
        conn.commit()
        return jsonify({"ok": True})
    except sqlite3.Error:
        conn.rollback()
        return jsonify({"ok": False, "error": "Database error"}), 500
    finally:
        conn.close()


@app.route("/admin/credit/<int:req_id>/reject", methods=["POST"])
@admin_required
def admin_credit_reject(req_id: int):
    note = (request.form.get("note") or "").strip() or None
    with get_db() as conn:
        cur = conn.execute(
            "UPDATE credit_requests SET status = 'rejected', reviewed_at = CURRENT_TIMESTAMP, review_note = ? WHERE id = ? AND status = 'pending'",
            (note, req_id),
        )
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({"ok": False, "error": "Request not found or already processed"}), 404
    return jsonify({"ok": True})


@app.route("/profile", methods=["GET", "POST"])
@student_or_teacher_required
def profile():
    rollno = session.get("rollno")
    if not rollno:
        return redirect(url_for("login"))
    user = get_user_by_rollno(rollno)
    if not user:
        return redirect(url_for("logout"))
    if request.method == "POST":
        # Profile saves may come as multipart/form-data (photo upload) or JSON.
        # Using silent=True prevents 415 errors for multipart requests.
        data = request.get_json(silent=True)
        if data is None:
            data = request.form
        tagline = (data.get("tagline") or "").strip()
        summary = (data.get("summary") or "").strip()
        # Categorized technical skills (stored as JSON)
        skills_prog = (data.get("skills_programming_languages") or "").strip()
        skills_web = (data.get("skills_web_technologies") or "").strip()
        skills_fw = (data.get("skills_frameworks_libraries") or "").strip()
        skills_tools = (data.get("skills_tools_platforms") or "").strip()
        skills_concepts = (data.get("skills_concepts") or "").strip()
        skills_obj = {
            "programming_languages": skills_prog,
            "web_technologies": skills_web,
            "frameworks_libraries": skills_fw,
            "tools_platforms": skills_tools,
            "concepts": skills_concepts,
        }
        skills_json = json.dumps(skills_obj, ensure_ascii=False)
        # Legacy: keep a combined comma string too (for any older UI bits)
        skills = ", ".join(
            [x for x in [skills_prog, skills_web, skills_fw, skills_tools, skills_concepts] if x]
        )
        non_tech_skills = (data.get("non_tech_skills") or "").strip()
        projects_raw = data.get("projects")
        if isinstance(projects_raw, str):
            projects = projects_raw
        else:
            projects = json.dumps(projects_raw) if projects_raw else "[]"
        achievements_raw = data.get("achievements")
        if isinstance(achievements_raw, str):
            achievements_json = achievements_raw
        else:
            achievements_json = json.dumps(achievements_raw or [])
        work_raw = data.get("work_experience")
        if isinstance(work_raw, str):
            work_experience_json = work_raw
        else:
            work_experience_json = json.dumps(work_raw or [])
        edu_raw = data.get("education")
        if isinstance(edu_raw, str):
            education_json = edu_raw
        else:
            education_json = json.dumps(edu_raw or [])
        location = (data.get("location") or "").strip()
        linkedin = (data.get("linkedin") or "").strip()
        github = (data.get("github") or "").strip()
        portfolio = (data.get("portfolio") or "").strip()
        email = (data.get("email") or "").strip()
        phone = (data.get("phone") or "").strip()
        is_public_raw = (data.get("is_public") or "").strip().lower()
        is_public = 1 if is_public_raw in ("1", "true", "yes", "on") else 0
        photo_url = None
        photo = request.files.get("photo")
        if photo and photo.filename:
            ext = (os.path.splitext(photo.filename)[1] or "").lower().lstrip(".")
            if ext in ALLOWED_PHOTO_EXT:
                safe_rollno = "".join(c for c in rollno if c.isalnum() or c in "._-") or "user"
                filename = secure_filename(f"{safe_rollno}.{ext}")
                os.makedirs(UPLOAD_DIR, exist_ok=True)
                filepath = os.path.join(UPLOAD_DIR, filename)
                if not os.path.abspath(filepath).startswith(os.path.abspath(UPLOAD_DIR)):
                    return jsonify({"ok": False, "error": "Invalid photo path"}), 400
                photo.save(filepath)
                photo_url = os.path.join("uploads", "profiles", filename).replace("\\", "/")
        with get_db() as conn:
            if not photo_url:
                row = conn.execute("SELECT photo_url FROM profiles WHERE rollno = ?", (rollno,)).fetchone()
                photo_url = row["photo_url"] if row and row["photo_url"] else ""
            cur = conn.execute("SELECT email, phone FROM users WHERE rollno = ?", (rollno,)).fetchone()
            prev_email, prev_phone = (cur["email"], cur["phone"]) if cur else ("", "")
            conn.execute("UPDATE users SET email = ?, phone = ? WHERE rollno = ?", (email or prev_email, phone or prev_phone, rollno))
            conn.execute(
                """INSERT INTO profiles (rollno, tagline, summary, skills, skills_json, non_tech_skills, projects, achievements_json, work_experience_json, education_json, location, linkedin, github, portfolio, is_public, photo_url, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                   ON CONFLICT(rollno) DO UPDATE SET tagline=excluded.tagline, summary=excluded.summary, skills=excluded.skills, skills_json=excluded.skills_json, non_tech_skills=excluded.non_tech_skills, projects=excluded.projects, achievements_json=excluded.achievements_json, work_experience_json=excluded.work_experience_json, education_json=excluded.education_json, location=excluded.location, linkedin=excluded.linkedin, github=excluded.github, portfolio=excluded.portfolio, is_public=excluded.is_public, photo_url=excluded.photo_url, updated_at=CURRENT_TIMESTAMP""",
                (
                    rollno,
                    tagline,
                    summary,
                    skills,
                    skills_json,
                    non_tech_skills,
                    projects,
                    achievements_json,
                    work_experience_json,
                    education_json,
                    location,
                    linkedin,
                    github,
                    portfolio,
                    is_public,
                    photo_url or None,
                ),
            )
            conn.commit()
        return jsonify({"ok": True})
    with get_db() as conn:
        try:
            pro = conn.execute(
                "SELECT tagline, summary, skills, skills_json, non_tech_skills, projects, achievements_json, work_experience_json, education_json, location, linkedin, github, portfolio, COALESCE(is_public, 0) AS is_public, photo_url FROM profiles WHERE rollno = ?",
                (rollno,),
            ).fetchone()
        except sqlite3.OperationalError:
            pro = conn.execute("SELECT tagline, summary, skills, projects, location, linkedin, github, portfolio, photo_url FROM profiles WHERE rollno = ?", (rollno,)).fetchone()
    profile_data = dict(pro) if pro else {}
    skills_json_raw = profile_data.get("skills_json") or ""
    skills_obj = {}
    if skills_json_raw:
        try:
            skills_obj = json.loads(skills_json_raw) or {}
        except (json.JSONDecodeError, TypeError):
            skills_obj = {}
    # Backward compatibility: if no skills_json yet, interpret legacy skills list as frameworks/tools.
    if not isinstance(skills_obj, dict) or not skills_obj:
        legacy = [s.strip() for s in (profile_data.get("skills") or "").split(",") if s.strip()]
        skills_obj = {
            "programming_languages": "",
            "web_technologies": "",
            "frameworks_libraries": ", ".join(legacy),
            "tools_platforms": "",
            "concepts": "",
        }
    non_tech_str = profile_data.get("non_tech_skills", "") or ""
    non_tech_skills_list = [s.strip() for s in non_tech_str.split(",") if s.strip()]
    try:
        projects_list = json.loads(profile_data.get("projects") or "[]")
    except (json.JSONDecodeError, TypeError):
        projects_list = []
    try:
        achievements_list = json.loads(profile_data.get("achievements_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        achievements_list = []
    try:
        work_experience_list = json.loads(profile_data.get("work_experience_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        work_experience_list = []
    try:
        education_list = json.loads(profile_data.get("education_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        education_list = []
    photo_url = profile_data.get("photo_url") or ""
    return render_template(
        "profile.html",
        name=user["name"],
        rollno=user["rollno"],
        email=user["email"],
        phone=user["phone"],
        department=user["department"],
        year=user["year"],
        tagline=profile_data.get("tagline") or "",
        summary=profile_data.get("summary") or "",
        skills_obj=skills_obj,
        non_tech_skills_list=non_tech_skills_list,
        projects_list=projects_list,
        achievements_list=achievements_list,
        work_experience_list=work_experience_list,
        education_list=education_list,
        location=profile_data.get("location") or "",
        linkedin=profile_data.get("linkedin") or "",
        github=profile_data.get("github") or "",
        portfolio=profile_data.get("portfolio") or "",
        photo_url=photo_url,
        is_public=int(profile_data.get("is_public") or 0),
    )


@app.route("/profile/print")
@student_or_teacher_required
def profile_print():
    """Printable resume-style profile for the logged-in user (even if not public)."""
    rollno = session.get("rollno")
    if not rollno:
        return redirect(url_for("login"))
    user = get_user_by_rollno(rollno)
    if not user:
        return redirect(url_for("logout"))
    with get_db() as conn:
        p = conn.execute(
            """SELECT tagline, summary, skills, non_tech_skills, projects, achievements_json, work_experience_json, education_json,
                      location, linkedin, github, portfolio, skills_json, COALESCE(is_public, 0) AS is_public, photo_url
               FROM profiles WHERE rollno = ?""",
            ((rollno or "").strip().upper(),),
        ).fetchone()
    profile_data = dict(p) if p else {}
    skills_obj = {}
    try:
        skills_obj = json.loads(profile_data.get("skills_json") or "") or {}
    except (json.JSONDecodeError, TypeError):
        skills_obj = {}
    if not isinstance(skills_obj, dict) or not skills_obj:
        legacy = [s.strip() for s in (profile_data.get("skills") or "").split(",") if s.strip()]
        skills_obj = {
            "programming_languages": "",
            "web_technologies": "",
            "frameworks_libraries": ", ".join(legacy),
            "tools_platforms": "",
            "concepts": "",
        }
    non_tech_skills_list = [s.strip() for s in (profile_data.get("non_tech_skills") or "").split(",") if s.strip()]
    try:
        projects_list = json.loads(profile_data.get("projects") or "[]")
    except (json.JSONDecodeError, TypeError):
        projects_list = []
    try:
        achievements_list = json.loads(profile_data.get("achievements_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        achievements_list = []
    try:
        work_experience_list = json.loads(profile_data.get("work_experience_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        work_experience_list = []
    try:
        education_list = json.loads(profile_data.get("education_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        education_list = []
    return render_template(
        "public_profile.html",
        name=user["name"],
        rollno=user["rollno"],
        department=user["department"],
        year=user["year"],
        tagline=profile_data.get("tagline") or "",
        summary=profile_data.get("summary") or "",
        skills_obj=skills_obj,
        non_tech_skills_list=non_tech_skills_list,
        projects_list=projects_list,
        achievements_list=achievements_list,
        work_experience_list=work_experience_list,
        education_list=education_list,
        location=profile_data.get("location") or "",
        linkedin=profile_data.get("linkedin") or "",
        github=profile_data.get("github") or "",
        portfolio=profile_data.get("portfolio") or "",
        photo_url=profile_data.get("photo_url") or "",
        print_mode=True,
    )


@app.route("/profile/download-pdf")
@student_or_teacher_required
def profile_download_pdf():
    """Direct PDF download for the logged-in user's profile."""
    rollno = session.get("rollno")
    if not rollno:
        return redirect(url_for("login"))
    user = get_user_by_rollno(rollno)
    if not user:
        return redirect(url_for("logout"))

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
        from PIL import Image, ImageDraw
    except Exception:
        return jsonify({"ok": False, "error": "PDF engine is unavailable. Install 'reportlab' and retry."}), 500

    with get_db() as conn:
        p = conn.execute(
            """SELECT tagline, summary, skills, non_tech_skills, projects, achievements_json, work_experience_json, education_json,
                      location, linkedin, github, portfolio, photo_url
               FROM profiles WHERE rollno = ?""",
            ((rollno or "").strip().upper(),),
        ).fetchone()
    profile_data = dict(p) if p else {}

    try:
        education_list = json.loads(profile_data.get("education_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        education_list = []
    try:
        work_experience_list = json.loads(profile_data.get("work_experience_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        work_experience_list = []
    try:
        projects_list = json.loads(profile_data.get("projects") or "[]")
    except (json.JSONDecodeError, TypeError):
        projects_list = []

    from reportlab.lib import colors

    bio = BytesIO()
    pdf = canvas.Canvas(bio, pagesize=A4)
    page_w, page_h = A4
    user_data = dict(user)

    # Resume-style page blocks (matching website theme)
    margin = 26
    sidebar_w = 170
    pdf.setFillColor(colors.HexColor("#0F172A"))
    pdf.rect(margin, margin, sidebar_w, page_h - (margin * 2), fill=1, stroke=0)
    pdf.setFillColor(colors.HexColor("#F8FAFC"))
    pdf.rect(margin + sidebar_w, margin, page_w - margin * 2 - sidebar_w, page_h - (margin * 2), fill=1, stroke=0)

    def wrap_text(txt: str, max_w: float, font_name: str, font_size: int) -> list[str]:
        words = (txt or "").split()
        if not words:
            return []
        out, cur = [], words[0]
        for w in words[1:]:
            cand = f"{cur} {w}"
            if pdf.stringWidth(cand, font_name, font_size) <= max_w:
                cur = cand
            else:
                out.append(cur)
                cur = w
        out.append(cur)
        return out

    def wrap_text_hard(txt: str, max_w: float, font_name: str, font_size: int) -> list[str]:
        """Wraps even long URL-like tokens with no spaces."""
        s = (txt or "").strip()
        if not s:
            return []
        lines: list[str] = []
        cur = ""
        for ch in s:
            cand = cur + ch
            if pdf.stringWidth(cand, font_name, font_size) <= max_w:
                cur = cand
            else:
                if cur:
                    lines.append(cur)
                cur = ch
        if cur:
            lines.append(cur)
        return lines

    # Left sidebar content
    sx = margin + 14
    sy = page_h - 34
    # Profile image block (to match resume template style)
    img_size = 86
    img_x = margin + (sidebar_w - img_size) / 2
    img_y = sy - img_size
    photo_rel = (profile_data.get("photo_url") or "").replace("\\", "/").strip("/")
    if photo_rel and ".." not in photo_rel:
        photo_abs = os.path.join(os.path.dirname(__file__), "static", photo_rel.replace("/", os.sep))
        if os.path.isfile(photo_abs):
            try:
                # Build a true circular image (alpha mask) before drawing.
                with Image.open(photo_abs) as im:
                    im = im.convert("RGBA")
                    side = min(im.size[0], im.size[1])
                    left_crop = (im.size[0] - side) // 2
                    top_crop = (im.size[1] - side) // 2
                    im = im.crop((left_crop, top_crop, left_crop + side, top_crop + side))
                    im = im.resize((int(img_size), int(img_size)), Image.Resampling.LANCZOS)

                    mask = Image.new("L", (int(img_size), int(img_size)), 0)
                    mdraw = ImageDraw.Draw(mask)
                    mdraw.ellipse((0, 0, int(img_size) - 1, int(img_size) - 1), fill=255)
                    im.putalpha(mask)

                    img_buf = BytesIO()
                    im.save(img_buf, format="PNG")
                    img_buf.seek(0)
                    pdf.drawImage(
                        ImageReader(img_buf),
                        img_x,
                        img_y,
                        width=img_size,
                        height=img_size,
                        preserveAspectRatio=True,
                        mask="auto",
                    )
            except Exception:
                pass
    pdf.setStrokeColor(colors.HexColor("#64748B"))
    pdf.setLineWidth(1.4)
    pdf.circle(margin + sidebar_w / 2, img_y + img_size / 2, img_size / 2 + 3, stroke=1, fill=0)
    sy = img_y - 14
    pdf.setFont("Helvetica", 9)
    pdf.setFillColor(colors.HexColor("#CBD5E1"))
    pdf.drawString(sx, sy, f"{user_data.get('department','')} | Year {user_data.get('year','')}")
    sy -= 24
    pdf.setFillColor(colors.HexColor("#A7F3D0"))
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(sx, sy, "CONTACT")
    sy -= 14
    pdf.setFillColor(colors.HexColor("#CBD5E1"))
    pdf.setFont("Helvetica", 9)
    for line in [
        str(user_data.get("rollno") or ""),
        str(profile_data.get("location") or ""),
        str(profile_data.get("linkedin") or ""),
        str(profile_data.get("github") or ""),
        str(profile_data.get("portfolio") or ""),
    ]:
        if line.strip():
            for ln in wrap_text_hard(line.strip(), sidebar_w - 24, "Helvetica", 9)[:3]:
                pdf.drawString(sx, sy, ln[:72])
                sy -= 11
    sy -= 10
    if education_list:
        pdf.setFillColor(colors.HexColor("#A7F3D0"))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(sx, sy, "EDUCATION")
        sy -= 16
        for ed in education_list[:3]:
            pdf.setFillColor(colors.HexColor("#F8FAFC"))
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(sx, sy, (ed.get("institution") or "")[:34])
            sy -= 10
            pdf.setFillColor(colors.HexColor("#CBD5E1"))
            pdf.setFont("Helvetica", 8)
            meta = " | ".join(
                x for x in [(ed.get("degree") or "").strip(), (ed.get("duration") or "").strip(), (ed.get("grade") or "").strip()] if x
            )
            for ln in wrap_text(meta, sidebar_w - 24, "Helvetica", 8)[:2]:
                pdf.drawString(sx, sy, ln[:44])
                sy -= 10
            sy -= 7

    skills_text = ", ".join(
        x
        for x in [
            (profile_data.get("skills") or "").strip(),
            (profile_data.get("non_tech_skills") or "").strip(),
        ]
        if x
    )
    if skills_text and sy > margin + 54:
        pdf.setFillColor(colors.HexColor("#A7F3D0"))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(sx, sy, "SKILLS")
        sy -= 16
        pdf.setFillColor(colors.HexColor("#CBD5E1"))
        pdf.setFont("Helvetica", 8.5)
        for ln in wrap_text_hard(skills_text, sidebar_w - 24, "Helvetica", 8.5)[:14]:
            if sy < margin + 18:
                break
            pdf.drawString(sx, sy, ln[:78])
            sy -= 10

    # Right panel content
    rx = margin + sidebar_w + 20
    ry = page_h - 58
    right_w = page_w - rx - margin - 8

    def draw_heading(title: str):
        nonlocal ry
        pdf.setFillColor(colors.HexColor("#0F172A"))
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(rx, ry, title.upper())
        ry -= 10
        pdf.setStrokeColor(colors.HexColor("#94A3B8"))
        pdf.line(rx, ry, rx + right_w, ry)
        ry -= 18

    def draw_para(txt: str, font_name: str = "Helvetica", font_size: int = 10, color_hex: str = "#334155", line_gap: int = 12):
        nonlocal ry
        pdf.setFillColor(colors.HexColor(color_hex))
        pdf.setFont(font_name, font_size)
        lines = wrap_text((txt or "").strip(), right_w, font_name, font_size)
        if not lines:
            return
        for ln in lines:
            if ry < 56:
                return
            pdf.drawString(rx, ry, ln[:150])
            ry -= line_gap

    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawString(rx, ry, (user_data.get("name") or "")[:28].upper())
    ry -= 18
    pdf.setFillColor(colors.HexColor("#334155"))
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(rx, ry, (profile_data.get("tagline") or "Student Profile")[:48])
    ry -= 24

    draw_heading("About me")
    draw_para((profile_data.get("summary") or "").strip() or "-", line_gap=13)
    ry -= 14

    if work_experience_list:
        draw_heading("Work experience")
        for w in work_experience_list[:5]:
            role = (w.get("role") or "").strip() or "Role"
            org = (w.get("org") or "").strip() or "Organization"
            meta = " • ".join(x for x in [(w.get("duration") or "").strip(), (w.get("location") or "").strip()] if x)
            left_w = right_w * 0.62
            right_w_meta = right_w * 0.34
            role_lines = wrap_text(f"{role} @ {org}", left_w, "Helvetica-Bold", 11)[:2]
            meta_lines = wrap_text(meta, right_w_meta, "Helvetica-Bold", 9)[:2] if meta else []
            row_lines = max(len(role_lines), len(meta_lines), 1)
            for i in range(row_lines):
                if ry < 72:
                    break
                if i < len(role_lines):
                    pdf.setFillColor(colors.HexColor("#111827"))
                    pdf.setFont("Helvetica-Bold", 11)
                    pdf.drawString(rx, ry, role_lines[i])
                if i < len(meta_lines):
                    pdf.setFillColor(colors.HexColor("#475569"))
                    pdf.setFont("Helvetica-Bold", 9)
                    pdf.drawRightString(rx + right_w, ry, meta_lines[i])
                ry -= 12
            draw_para((w.get("desc") or "").strip(), font_size=9, color_hex="#334155", line_gap=11)
            ry -= 12
            if ry < 72:
                break

    if ry > 130 and projects_list:
        draw_heading("Projects")
        for p in projects_list[:4]:
            title = (p.get("title") or "").strip() or "Project"
            tech = (p.get("tech") or "").strip()
            desc = (p.get("desc") or "").strip()
            t_lines = wrap_text(title, right_w * 0.62, "Helvetica-Bold", 11)[:2]
            m_lines = wrap_text(f"Tech: {tech}", right_w * 0.34, "Helvetica-Bold", 9)[:2] if tech else []
            row_lines = max(len(t_lines), len(m_lines), 1)
            for i in range(row_lines):
                if ry < 72:
                    break
                if i < len(t_lines):
                    pdf.setFillColor(colors.HexColor("#111827"))
                    pdf.setFont("Helvetica-Bold", 11)
                    pdf.drawString(rx, ry, t_lines[i])
                if i < len(m_lines):
                    pdf.setFillColor(colors.HexColor("#475569"))
                    pdf.setFont("Helvetica-Bold", 9)
                    pdf.drawRightString(rx + right_w, ry, m_lines[i])
                ry -= 12
            draw_para(desc, font_size=9, color_hex="#334155", line_gap=11)
            ry -= 12
            if ry < 72:
                break

    pdf.save()
    bio.seek(0)
    out_name = secure_filename(f"{(user_data.get('rollno') or 'profile').lower()}_profile.pdf") or "profile.pdf"
    return send_file(
        bio,
        as_attachment=True,
        download_name=out_name,
        mimetype="application/pdf",
    )


@app.route("/p/<rollno>")
def public_profile(rollno: str):
    rn = (rollno or "").strip().upper()
    if not rn:
        return redirect(url_for("landing"))
    with get_db() as conn:
        u = conn.execute(
            "SELECT rollno, name, department, year FROM users WHERE rollno = ?",
            (rn,),
        ).fetchone()
        if not u:
            return redirect(url_for("landing"))
        p = conn.execute(
            """SELECT tagline, summary, skills, non_tech_skills, projects, achievements_json, work_experience_json, education_json,
                      location, linkedin, github, portfolio, skills_json, COALESCE(is_public, 0) AS is_public, photo_url
               FROM profiles WHERE rollno = ?""",
            (rn,),
        ).fetchone()
    if not p or int(p["is_public"] or 0) != 1:
        return redirect(url_for("landing"))
    print_mode = (request.args.get("print") or "").strip().lower() in ("1", "true", "yes", "on")
    profile_data = dict(p)
    skills_obj = {}
    try:
        skills_obj = json.loads(profile_data.get("skills_json") or "") or {}
    except (json.JSONDecodeError, TypeError):
        skills_obj = {}
    if not isinstance(skills_obj, dict) or not skills_obj:
        legacy = [s.strip() for s in (profile_data.get("skills") or "").split(",") if s.strip()]
        skills_obj = {
            "programming_languages": "",
            "web_technologies": "",
            "frameworks_libraries": ", ".join(legacy),
            "tools_platforms": "",
            "concepts": "",
        }
    non_tech_skills_list = [s.strip() for s in (profile_data.get("non_tech_skills") or "").split(",") if s.strip()]
    try:
        projects_list = json.loads(profile_data.get("projects") or "[]")
    except (json.JSONDecodeError, TypeError):
        projects_list = []
    try:
        achievements_list = json.loads(profile_data.get("achievements_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        achievements_list = []
    try:
        work_experience_list = json.loads(profile_data.get("work_experience_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        work_experience_list = []
    try:
        education_list = json.loads(profile_data.get("education_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        education_list = []
    return render_template(
        "public_profile.html",
        name=u["name"],
        rollno=u["rollno"],
        department=u["department"],
        year=u["year"],
        tagline=profile_data.get("tagline") or "",
        summary=profile_data.get("summary") or "",
        skills_obj=skills_obj,
        non_tech_skills_list=non_tech_skills_list,
        projects_list=projects_list,
        achievements_list=achievements_list,
        work_experience_list=work_experience_list,
        education_list=education_list,
        location=profile_data.get("location") or "",
        linkedin=profile_data.get("linkedin") or "",
        github=profile_data.get("github") or "",
        portfolio=profile_data.get("portfolio") or "",
        photo_url=profile_data.get("photo_url") or "",
        print_mode=print_mode,
    )


@app.route("/payment")
def payment():
    """Shared NFC terminal: pay TO a roll number; payer is identified by card tap + PIN."""
    name = rollno = None
    wallet_balance = None
    if session.get("rollno"):
        user = get_user_by_rollno(session["rollno"])
        if user:
            name = user["name"]
            rollno = user["rollno"]
            wallet_balance = int(user["wallet_balance"] or 0)
    return render_template(
        "payment.html",
        name=name,
        rollno=rollno,
        wallet_balance=wallet_balance,
        logged_in=bool(session.get("user_id")),
    )


@app.route("/attendance")
@teacher_or_admin_required
def attendance_page():
    user = get_user_by_rollno(session.get("rollno")) if session.get("rollno") else None
    r = session_role()
    rollno = session.get("rollno") or ""
    dashboard_href = (
        url_for("admin_dashboard") if r == ROLE_ADMIN else url_for("teacher_dashboard")
    )
    with get_db() as conn:
        ac_rows = conn.execute(
            """SELECT name_key, display_name, COALESCE(semester_no, 1) AS semester_no
               FROM attendance_classes ORDER BY display_name COLLATE NOCASE"""
        ).fetchall()
        attendance_subjects = _attendance_subject_options(conn)
    attendance_classes = [
        {
            "name_key": x["name_key"],
            "display_name": x["display_name"],
            "semester_no": int(x["semester_no"] or 1),
        }
        for x in ac_rows
    ]
    return render_template(
        "attendance.html",
        name=(user["name"] if user else session.get("name")) or "User",
        dashboard_href=dashboard_href,
        attendance_classes=attendance_classes,
        attendance_subjects=attendance_subjects,
        session_date_default=date.today().isoformat(),
    )


@app.route("/api/attendance-session-check", methods=["GET"])
@teacher_or_admin_required
def api_attendance_session_check():
    class_key = (request.args.get("class_key") or "").strip()
    period = (request.args.get("period") or "").strip()
    session_date = (request.args.get("session_date") or "").strip()
    if not class_key or not period or not session_date:
        return jsonify({"ok": False, "error": "class_key, period, and session_date required"}), 400
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", session_date):
        return jsonify({"ok": False, "error": "Invalid session_date"}), 400
    teacher_roll = (session.get("rollno") or "").strip().upper()
    with get_db() as conn:
        known_class = conn.execute("SELECT 1 FROM attendance_classes WHERE name_key = ?", (class_key,)).fetchone()
        # Use Excel as the source of truth for "taken" (matches the Period dropdown behavior).
        excel_periods = _attendance_excel_periods(class_key, session_date)
        taken_by_excel = False
        for p in excel_periods or []:
            if str(p.get("period") or "").strip() == str(period).strip():
                taken_by_excel = bool(p.get("taken"))
                break
        # If DB has an existing session, only block if it was submitted by someone else.
        existing = conn.execute(
            """SELECT teacher_rollno
               FROM attendance_sessions
               WHERE class_key = ? AND period = ? AND session_date = ?""",
            (class_key, period, session_date),
        ).fetchone()
    if not known_class:
        return jsonify({"ok": False, "error": "Unknown class", "taken": False}), 400
    if taken_by_excel:
        return jsonify(
            {
                "ok": True,
                "taken": True,
                "message": "Attendance for this class/period is already marked as taken in the Excel sheet.",
            }
        )
    if existing:
        prev_teacher = (existing["teacher_rollno"] or "").strip().upper()
        if prev_teacher and teacher_roll and prev_teacher != teacher_roll:
            return jsonify(
                {
                    "ok": True,
                    "taken": True,
                    "message": "Attendance for this class, period, and date was already submitted by another teacher.",
                }
            )
    return jsonify({"ok": True, "taken": False})


@app.route("/api/attendance-period-options", methods=["GET"])
@teacher_or_admin_required
def api_attendance_period_options():
    class_key = (request.args.get("class_key") or "").strip()
    session_date = (request.args.get("session_date") or "").strip()
    if not class_key or not session_date:
        return jsonify({"ok": False, "error": "class_key and session_date required"}), 400
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", session_date):
        return jsonify({"ok": False, "error": "Invalid session_date"}), 400
    periods = _attendance_excel_periods(class_key, session_date)
    if not periods:
        # Fallback shape: no date sheet yet -> all periods open.
        periods = [{"period": str(p), "label": str(p), "taken": False} for p in range(1, 10)]
    return jsonify({"ok": True, "periods": periods})


@app.route("/api/attendance-lookup", methods=["POST"])
@teacher_or_admin_required
def api_attendance_lookup():
    """Resolve NFC / reader UID to registered student (cards table)."""
    data = request.get_json() or {}
    card_uid = (data.get("card_uid") or "").strip()
    if not card_uid:
        return jsonify({"ok": False, "error": "No card UID"}), 400
    norm = normalize_card_uid(card_uid)
    if not norm:
        return jsonify({"ok": False, "error": "Invalid UID"}), 400
    with get_db() as conn:
        row = conn.execute(
            """SELECT c.rollno, u.name, c.card_uid FROM cards c
               JOIN users u ON u.rollno = c.rollno
               WHERE upper(replace(replace(replace(trim(c.card_uid), '-', ''), ':', ''), ' ', '')) = ?""",
            (norm,),
        ).fetchone()
    if not row:
        return jsonify({"ok": False, "error": "Card not registered", "card_uid": card_uid}), 404
    return jsonify(
        {
            "ok": True,
            "rollno": row["rollno"],
            "name": row["name"],
            "card_uid": row["card_uid"] or card_uid,
        }
    )


@app.route("/api/attendance-submit", methods=["POST"])
@teacher_or_admin_required
def api_attendance_submit():
    """Persist session + rows; one class+period+date per day (locked after first submit)."""
    data = request.get_json() or {}
    teacher = session.get("rollno") or ""
    entries = data.get("entries")
    if not isinstance(entries, list) or not entries:
        return jsonify({"ok": False, "error": "No entries to save"}), 400
    class_key = (data.get("class_key") or "").strip()
    period = (data.get("period") or "").strip()
    periods_raw = data.get("periods")
    periods: list[str]
    if isinstance(periods_raw, list) and periods_raw:
        periods = [str(p).strip() for p in periods_raw if str(p).strip()]
    elif period:
        periods = [period]
    else:
        periods = []
    subject_key = (data.get("subject_key") or "").strip()
    session_date = (data.get("session_date") or "").strip()
    if not class_key or not periods:
        return jsonify({"ok": False, "error": "class_key and period(s) required"}), 400
    if not session_date:
        session_date = date.today().isoformat()
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", session_date):
        return jsonify({"ok": False, "error": "Invalid session_date"}), 400
    # Validate periods are 1..9
    clean_periods: list[str] = []
    for p in periods:
        pp = str(p).strip()
        if not pp:
            continue
        if not re.fullmatch(r"\d+", pp):
            return jsonify({"ok": False, "error": "Invalid period value"}), 400
        n = int(pp)
        if not 1 <= n <= 9:
            return jsonify({"ok": False, "error": "Period must be between 1 and 9"}), 400
        clean_periods.append(str(n))
    # de-dupe preserving order
    seen_p = set()
    clean_periods = [p for p in clean_periods if not (p in seen_p or seen_p.add(p))]
    if not clean_periods:
        return jsonify({"ok": False, "error": "No valid periods provided"}), 400
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    saved = 0
    excel_url = None
    try:
        conn.execute("BEGIN IMMEDIATE")
        ck = conn.execute(
            "SELECT 1 FROM attendance_classes WHERE name_key = ?",
            (class_key,),
        ).fetchone()
        if not ck:
            conn.rollback()
            return jsonify({"ok": False, "error": "Class is not defined. Ask admin to create it."}), 400
        # Insert one session per period (continuous periods supported)
        for p in clean_periods:
            try:
                conn.execute(
                    """INSERT INTO attendance_sessions (teacher_rollno, class_key, period, session_date, subject_key)
                       VALUES (?, ?, ?, ?, ?)""",
                    (teacher, class_key, p, session_date, subject_key or None),
                )
            except sqlite3.IntegrityError:
                # If a previous session exists for same teacher but Excel shows it's not taken,
                # allow overwrite (helps recover from partial failures where DB was written but Excel wasn't).
                existing = conn.execute(
                    """SELECT teacher_rollno
                       FROM attendance_sessions
                       WHERE class_key = ? AND period = ? AND session_date = ?""",
                    (class_key, p, session_date),
                ).fetchone()
                prev_teacher = (existing["teacher_rollno"] or "").strip().upper() if existing else ""
                if prev_teacher and prev_teacher != (teacher or "").strip().upper():
                    conn.rollback()
                    return jsonify(
                        {
                            "ok": False,
                            "error": f"Attendance for period {p} was already submitted by another teacher.",
                        }
                    ), 409
                excel_periods = _attendance_excel_periods(class_key, session_date)
                taken_by_excel = False
                for ep in excel_periods or []:
                    if str(ep.get("period") or "").strip() == str(p).strip():
                        taken_by_excel = bool(ep.get("taken"))
                        break
                if taken_by_excel:
                    conn.rollback()
                    return jsonify(
                        {
                            "ok": False,
                            "error": f"Attendance for period {p} is already marked taken in Excel.",
                        }
                    ), 409
                # Overwrite previous DB session + records for this teacher/period/date.
                conn.execute(
                    """DELETE FROM attendance_records
                       WHERE class_key = ? AND period = ? AND session_date = ? AND teacher_rollno = ?""",
                    (class_key, p, session_date, teacher),
                )
                conn.execute(
                    """DELETE FROM attendance_sessions
                       WHERE class_key = ? AND period = ? AND session_date = ? AND teacher_rollno = ?""",
                    (class_key, p, session_date, teacher),
                )
                conn.execute(
                    """INSERT INTO attendance_sessions (teacher_rollno, class_key, period, session_date, subject_key)
                       VALUES (?, ?, ?, ?, ?)""",
                    (teacher, class_key, p, session_date, subject_key or None),
                )

        rows_for_excel: list[tuple[str, str]] = []
        seen_excel: set[str] = set()
        for e in entries:
            if not isinstance(e, dict):
                continue
            roll = (e.get("rollno") or "").strip().upper()
            uid = (e.get("card_uid") or "").strip()
            if not roll:
                continue
            for p in clean_periods:
                conn.execute(
                    """INSERT INTO attendance_records
                       (teacher_rollno, class_key, period, subject_key, student_rollno, card_uid, session_date)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        teacher,
                        class_key,
                        p,
                        subject_key or None,
                        roll,
                        uid or None,
                        session_date,
                    ),
                )
                saved += 1
            if roll not in seen_excel:
                seen_excel.add(roll)
                urow = conn.execute("SELECT name FROM users WHERE rollno = ?", (roll,)).fetchone()
                nm = (urow["name"] if urow else "") or ""
                rows_for_excel.append((roll, nm))
        conn.commit()
        try:
            # For continuous periods, write the same attendance into each period column.
            for p in clean_periods:
                write_attendance_session_excel(
                    class_key, p, session_date, subject_key or None, rows_for_excel
                )
        except Exception:
            app.logger.exception("attendance_session_excel")
    except sqlite3.Error:
        conn.rollback()
        return jsonify({"ok": False, "error": "Database error"}), 500
    finally:
        conn.close()
    payload = {"ok": True, "saved": saved}
    if excel_url:
        payload["excel_url"] = excel_url
    return jsonify(payload)


@app.route("/api/attendance-edit-save", methods=["POST"])
@teacher_or_admin_required
def api_attendance_edit_save():
    """Save edited P/A marks for one class on one date (for already created sessions)."""
    data = request.get_json() or {}
    class_key = (data.get("class_key") or "").strip()
    session_date = (data.get("session_date") or "").strip()
    rows = data.get("rows")
    if not class_key:
        return jsonify({"ok": False, "error": "class_key is required"}), 400
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", session_date):
        return jsonify({"ok": False, "error": "Valid session_date is required"}), 400
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "rows must be a list"}), 400

    role = session_role()
    teacher_roll = (session.get("rollno") or "").strip()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("BEGIN IMMEDIATE")
        cls = conn.execute(
            """SELECT roll_prefix, start_seq, end_seq, pad_width, missing_json
               FROM attendance_classes WHERE name_key = ?""",
            (class_key,),
        ).fetchone()
        if not cls:
            conn.rollback()
            return jsonify({"ok": False, "error": "Class not found"}), 404

        missing = parse_missing_numbers(cls["missing_json"] or "")
        all_rolls = build_roll_numbers(
            cls["roll_prefix"],
            int(cls["start_seq"]),
            int(cls["end_seq"]),
            missing,
            int(cls["pad_width"] or 0),
        )
        allowed_rolls = {r.upper() for r in all_rolls}

        sess_sql = """SELECT period, teacher_rollno, subject_key
                      FROM attendance_sessions
                      WHERE class_key = ? AND session_date = ?"""
        sess_params: list[str] = [class_key, session_date]
        if role == ROLE_TEACHER:
            sess_sql += " AND teacher_rollno = ?"
            sess_params.append(teacher_roll)
        sessions = conn.execute(sess_sql, sess_params).fetchall()
        if not sessions:
            conn.rollback()
            return jsonify({"ok": False, "error": "No attendance sessions found for this class/date"}), 400

        period_meta: dict[str, tuple[str, str | None]] = {}
        for s in sessions:
            p = (s["period"] or "").strip()
            if not p:
                continue
            period_meta[p] = ((s["teacher_rollno"] or "").strip(), s["subject_key"])
        if not period_meta:
            conn.rollback()
            return jsonify({"ok": False, "error": "No editable periods found"}), 400

        # Clear existing records for targeted sessions first.
        for p, (teacher_for_period, _) in period_meta.items():
            conn.execute(
                """DELETE FROM attendance_records
                   WHERE class_key = ? AND period = ? AND session_date = ? AND teacher_rollno = ?""",
                (class_key, p, session_date, teacher_for_period),
            )

        mark_map: dict[str, dict[str, str]] = {}
        for item in rows:
            if not isinstance(item, dict):
                continue
            roll = (item.get("rollno") or "").strip().upper()
            marks = item.get("marks")
            if not roll or roll not in allowed_rolls or not isinstance(marks, dict):
                continue
            clean_marks: dict[str, str] = {}
            for p, m in marks.items():
                pp = str(p).strip()
                mm = str(m).strip().upper()
                if pp in period_meta and mm in ("P", "A"):
                    clean_marks[pp] = mm
            if clean_marks:
                mark_map[roll] = clean_marks

        inserted = 0
        for roll, marks in mark_map.items():
            for p, m in marks.items():
                if m != "P":
                    continue
                teacher_for_period, subject_for_period = period_meta[p]
                conn.execute(
                    """INSERT INTO attendance_records
                       (teacher_rollno, class_key, period, subject_key, student_rollno, card_uid, session_date)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        teacher_for_period,
                        class_key,
                        p,
                        subject_for_period,
                        roll,
                        None,
                        session_date,
                    ),
                )
                inserted += 1

        conn.commit()
        return jsonify({"ok": True, "saved_present_records": inserted, "periods": sorted(period_meta.keys())})
    except sqlite3.Error:
        conn.rollback()
        return jsonify({"ok": False, "error": "Database error"}), 500
    finally:
        conn.close()


def _attendance_visibility_clause(
    role: str,
    rollno: str,
    incharge_classes: set[str] | None = None,
    assigned_classes: set[str] | None = None,
    assigned_subjects: set[str] | None = None,
):
    if role == ROLE_STUDENT:
        return "ar.student_rollno = ?", [rollno]
    if role == ROLE_TEACHER:
        incharge_classes = incharge_classes or set()
        assigned_classes = assigned_classes or set()
        assigned_subjects = assigned_subjects or set()
        conds = ["ar.teacher_rollno = ?"]
        params: list[str] = [rollno]
        if incharge_classes:
            ph = ",".join("?" for _ in sorted(incharge_classes))
            conds.append(f"ar.class_key IN ({ph})")
            params.extend(sorted(incharge_classes))
        if assigned_classes and assigned_subjects:
            phc = ",".join("?" for _ in sorted(assigned_classes))
            phs = ",".join("?" for _ in sorted(assigned_subjects))
            conds.append(f"(ar.class_key IN ({phc}) AND ar.subject_key IN ({phs}))")
            params.extend(sorted(assigned_classes))
            params.extend(sorted(assigned_subjects))
        return "(" + " OR ".join(conds) + ")", params
    return "1=1", []


_ATT_AR_DAY = "date(COALESCE(NULLIF(trim(ar.session_date), ''), ar.created_at))"


class ViewAttnQ(NamedTuple):
    start_date_q: str
    end_date_q: str
    class_q: str
    subject_q: str
    teacher_selected_incharge: bool
    allow_full_class_view: bool
    subject_date_view: bool
    non_incharge_scope_only: bool
    vis_sql: str
    vis_params: tuple


def _view_attendance_parse_qctx(
    r: str,
    rollno: str,
    teacher_incharge_classes: set[str],
    teacher_assigned_classes: set[str],
    teacher_assigned_subjects: set[str],
    args,
) -> ViewAttnQ:
    start_date_q = (args.get("start_date") or "").strip()
    end_date_q = (args.get("end_date") or "").strip()
    class_q = (args.get("class_key") or "").strip()
    subject_q = (args.get("subject_key") or "").strip()
    if class_q and not start_date_q and not end_date_q:
        today = date.today().isoformat()
        start_date_q = today
        end_date_q = today
    teacher_selected_incharge = bool(r == ROLE_TEACHER and class_q and class_q in teacher_incharge_classes)
    if r == ROLE_TEACHER and teacher_selected_incharge and not subject_q:
        subject_q = ""
    allow_full_class_view = bool(r == ROLE_ADMIN or teacher_selected_incharge)
    subject_date_view = bool(class_q and subject_q)
    non_incharge_scope_only = bool(r == ROLE_TEACHER and not teacher_selected_incharge)
    vis_sql, vis_params = _attendance_visibility_clause(
        r, rollno, teacher_incharge_classes, teacher_assigned_classes, teacher_assigned_subjects
    )
    return ViewAttnQ(
        start_date_q=start_date_q,
        end_date_q=end_date_q,
        class_q=class_q,
        subject_q=subject_q,
        teacher_selected_incharge=teacher_selected_incharge,
        allow_full_class_view=allow_full_class_view,
        subject_date_view=subject_date_view,
        non_incharge_scope_only=non_incharge_scope_only,
        vis_sql=vis_sql,
        vis_params=tuple(vis_params),
    )


def _va_base_conds(
    qctx: ViewAttnQ,
    teacher_assigned_classes: set[str],
    teacher_assigned_subjects: set[str],
) -> tuple[list[str], list]:
    conds = [f"({qctx.vis_sql})"]
    params: list = list(qctx.vis_params)
    if qctx.start_date_q:
        conds.append(f"{_ATT_AR_DAY} >= date(?)")
        params.append(qctx.start_date_q)
    if qctx.end_date_q:
        conds.append(f"{_ATT_AR_DAY} <= date(?)")
        params.append(qctx.end_date_q)
    if qctx.class_q:
        conds.append("ar.class_key = ?")
        params.append(qctx.class_q)
    if qctx.subject_q:
        conds.append("ar.subject_key = ?")
        params.append(qctx.subject_q)
    if qctx.non_incharge_scope_only:
        if (
            not qctx.class_q
            or not qctx.subject_q
            or qctx.class_q not in teacher_assigned_classes
            or qctx.subject_q not in teacher_assigned_subjects
        ):
            conds.append("1 = 0")
    return conds, params


def _view_attendance_auth():
    """Returns (redirect_response_or_None, role, rollno, password_only_admin, incharge, assigned_c, assigned_s)."""
    password_only_admin = bool(session.get("admin_ok")) and not bool(session.get("user_id"))
    if not session.get("user_id") and not password_only_admin:
        return redirect(url_for("login")), None, None, None, None, None, None
    if password_only_admin:
        r = ROLE_ADMIN
        rollno = ""
        incharge: set[str] = set()
        assigned_c: set[str] = set()
        assigned_s: set[str] = set()
    else:
        r = session_role()
        rollno = session.get("rollno") or ""
        if r not in (ROLE_STUDENT, ROLE_TEACHER, ROLE_ADMIN):
            return redirect_home_for_session(), None, None, None, None, None, None
        incharge, assigned_c, assigned_s = set(), set(), set()
        if r == ROLE_TEACHER and rollno:
            with get_db() as conn:
                incharge, assigned_c, assigned_s = _teacher_assignments(conn, rollno)
    return None, r, rollno, password_only_admin, incharge, assigned_c, assigned_s


def _safe_class_workbook_abs_path(conn: sqlite3.Connection, class_key: str) -> str | None:
    if not class_key:
        return None
    row = conn.execute("SELECT excel_path FROM attendance_classes WHERE name_key = ?", (class_key,)).fetchone()
    if not row or not (row["excel_path"] or "").strip():
        return None
    rel = (row["excel_path"] or "").replace("\\", "/").strip("/")
    if not rel or ".." in rel:
        return None
    abs_path = os.path.join(os.path.dirname(__file__), "static", rel.replace("/", os.sep))
    if not os.path.isfile(abs_path):
        return None
    return abs_path


def _workbook_bytes_xlsx(wb) -> BytesIO:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _export_period_norm_key(period: str) -> str:
    """Normalize P3 / 3 / period 3 → comparable key (digit or lowercase token)."""
    p = (period or "").strip()
    m = re.fullmatch(r"P?(\d+)$", p, re.I)
    if m:
        return m.group(1)
    return p.lower()


def _export_session_column_header(day: str, period_raw: str, pnorm: str) -> str:
    d = (day or "").strip()[:10]
    if pnorm.isdigit():
        return f"{d} P{pnorm}"
    return f"{d} {(period_raw or '').strip() or pnorm}"


def _export_entire_attendance_subject_date_columns(
    conn: sqlite3.Connection,
    where: str,
    params: list,
    wb,
    *,
    max_sheets: int = 100,
) -> None:
    """One worksheet per (class, subject): Roll, name, one column per session (date+period), total present, %."""
    if wb is None:
        raise ValueError("workbook required")
    lim = int(max_sheets)
    pairs = conn.execute(
        f"""SELECT DISTINCT ar.class_key, ar.subject_key
            FROM attendance_records ar
            WHERE ({where})
              AND ar.class_key IS NOT NULL AND trim(ar.class_key) != ''
              AND ar.subject_key IS NOT NULL AND trim(ar.subject_key) != ''
            ORDER BY ar.class_key COLLATE NOCASE, ar.subject_key COLLATE NOCASE
            LIMIT {lim}""",
        params,
    ).fetchall()
    if not pairs:
        ws = wb.create_sheet(title=_safe_excel_sheet_title("No data"))
        ws.append(["No attendance records match this export."])
        return

    used_titles: set[str] = set()
    day_expr = "date(COALESCE(NULLIF(trim(ar.session_date), ''), ar.created_at))"

    for pr in pairs:
        ck = (pr["class_key"] or "").strip()
        sk = (pr["subject_key"] or "").strip()
        if not ck or not sk:
            continue
        w_pair = f"({where}) AND ar.class_key = ? AND ar.subject_key = ?"
        p_pair = list(params) + [ck, sk]

        # One column per (date, period); same calendar day with P1,P2,P3 → three columns.
        col_order: list[tuple[str, str]] = []
        col_header: dict[tuple[str, str], str] = {}

        def _add_col(d_str: str, period_raw: str) -> None:
            d_str = (d_str or "").strip()[:10]
            praw = (period_raw or "").strip()
            if not d_str or not praw:
                return
            pn = _export_period_norm_key(praw)
            key = (d_str, pn)
            if key not in col_header:
                col_order.append(key)
                col_header[key] = _export_session_column_header(d_str, praw, pn)

        for rw in conn.execute(
            f"""SELECT DISTINCT {day_expr} AS d, trim(ar.period) AS period
                FROM attendance_records ar
                WHERE {w_pair}
                  AND ar.student_rollno IS NOT NULL AND trim(ar.student_rollno) != ''
                  AND ar.period IS NOT NULL AND trim(ar.period) != ''""",
            p_pair,
        ).fetchall():
            _add_col(str(rw["d"] or ""), str(rw["period"] or ""))

        for rw in conn.execute(
            """SELECT DISTINCT date(session_date) AS d, trim(period) AS period
               FROM attendance_sessions
               WHERE class_key = ? AND subject_key = ?
                 AND session_date IS NOT NULL AND trim(session_date) != ''
                 AND period IS NOT NULL AND trim(period) != ''""",
            (ck, sk),
        ).fetchall():
            _add_col(str(rw["d"] or ""), str(rw["period"] or ""))

        if not col_order:
            continue

        col_order.sort(
            key=lambda k: (
                k[0],
                int(k[1]) if k[1].isdigit() else 999,
                k[1],
            )
        )

        present: set = set()
        for rw in conn.execute(
            f"""SELECT DISTINCT upper(trim(ar.student_rollno)) AS roll,
                       {day_expr} AS d, trim(ar.period) AS period
                FROM attendance_records ar
                WHERE {w_pair}
                  AND ar.student_rollno IS NOT NULL AND trim(ar.student_rollno) != ''
                  AND ar.period IS NOT NULL AND trim(ar.period) != ''""",
            p_pair,
        ).fetchall():
            rl = (rw["roll"] or "").strip().upper()
            d = str(rw["d"] or "").strip()[:10]
            pr = str(rw["period"] or "").strip()
            if rl and d and pr:
                present.add((rl, d, _export_period_norm_key(pr)))

        selected_class = conn.execute(
            """SELECT roll_prefix, start_seq, end_seq, pad_width, missing_json
               FROM attendance_classes WHERE name_key = ?""",
            (ck,),
        ).fetchone()
        if selected_class:
            missing = parse_missing_numbers(selected_class["missing_json"] or "")
            all_rolls = build_roll_numbers(
                selected_class["roll_prefix"],
                int(selected_class["start_seq"]),
                int(selected_class["end_seq"]),
                missing,
                int(selected_class["pad_width"] or 0),
            )
        else:
            r_only = conn.execute(
                f"""SELECT DISTINCT trim(ar.student_rollno) AS roll
                    FROM attendance_records ar
                    WHERE {w_pair}
                      AND ar.student_rollno IS NOT NULL AND trim(ar.student_rollno) != ''""",
                p_pair,
            ).fetchall()
            all_rolls = sorted({(rw["roll"] or "").strip() for rw in r_only if (rw["roll"] or "").strip()}, key=str.upper)

        roll_names: dict[str, str] = {}
        if all_rolls:
            placeholders = ",".join("?" for _ in all_rolls)
            for rw in conn.execute(
                f"SELECT rollno, name FROM users WHERE rollno IN ({placeholders})",
                list(all_rolls),
            ).fetchall():
                roll_names[(rw["rollno"] or "").strip().upper()] = (rw["name"] or "").strip()

        sub_label = (SUBJECT_LABELS.get(sk, sk) or sk)[:18]
        base_title = _safe_excel_sheet_title(f"{ck} {sub_label}")[:31]
        st = base_title
        n = 2
        while st in used_titles:
            st = _safe_excel_sheet_title(f"{ck[:12]}_{sk}_{n}")[:31]
            n += 1
        used_titles.add(st)
        ws = wb.create_sheet(title=st)
        labels = [col_header[k] for k in col_order]
        num_sessions = len(col_order)
        ws.append(["Roll number", "Student name"] + labels + ["Total present", "Attendance %"])
        for roll in sorted(all_rolls, key=str.upper):
            ru = roll.strip().upper()
            marks = ["P" if (ru, d, pn) in present else "A" for d, pn in col_order]
            total_p = sum(1 for m in marks if m == "P")
            pct = round(100.0 * total_p / num_sessions, 2) if num_sessions else 0.0
            ws.append([roll, roll_names.get(ru, "")] + marks + [total_p, pct])

    if not wb.sheetnames:
        ws = wb.create_sheet(title=_safe_excel_sheet_title("No data"))
        ws.append(["No dated attendance rows matched this export."])
    elif len(pairs) >= lim:
        note = wb.create_sheet(title=_safe_excel_sheet_title("Note"))
        note.append(
            [
                f"At most {lim} class/subject sheets per file; more may exist in the database. Export again with narrower scope if needed."
            ]
        )


def _va_fill_period_grid_sheet(
    conn: sqlite3.Connection,
    ws,
    where: str,
    params: list,
    qctx: ViewAttnQ,
    r: str,
    rollno: str,
    header_date: str,
    export_limit: int,
) -> None:
    """One worksheet: period columns P1..P9 (or subject labels for single header_date)."""
    records = conn.execute(
        f"""SELECT ar.student_rollno,
                   COALESCE(u.name, '') AS student_name,
                   ar.period,
                   COUNT(*) AS present_count
            FROM attendance_records ar
            LEFT JOIN users u ON u.rollno = ar.student_rollno
            WHERE {where}
              AND ar.student_rollno IS NOT NULL
              AND trim(ar.student_rollno) != ''
            GROUP BY ar.student_rollno, u.name, ar.period
            ORDER BY ar.student_rollno COLLATE NOCASE
            LIMIT {export_limit}""",
        params,
    ).fetchall()

    class_q = qctx.class_q
    allow_full_class_view = qctx.allow_full_class_view
    subject_date_view = qctx.subject_date_view
    subject_q = qctx.subject_q
    start_date_q = qctx.start_date_q
    end_date_q = qctx.end_date_q

    selected_class = None
    taken_periods: set[str] = set()
    if class_q and (allow_full_class_view or subject_date_view):
        selected_class = conn.execute(
            """SELECT name_key, roll_prefix, start_seq, end_seq, pad_width, missing_json
               FROM attendance_classes WHERE name_key = ?""",
            (class_q,),
        ).fetchone()
        sess_conds = ["class_key = ?"]
        sess_params: list[str] = [class_q]
        if start_date_q:
            sess_conds.append("date(session_date) >= date(?)")
            sess_params.append(start_date_q)
        if end_date_q:
            sess_conds.append("date(session_date) <= date(?)")
            sess_params.append(end_date_q)
        if subject_q:
            sess_conds.append("subject_key = ?")
            sess_params.append(subject_q)
        if header_date:
            sess_conds.append("date(session_date) = date(?)")
            sess_params.append(header_date[:10])
        if r == ROLE_TEACHER and not allow_full_class_view:
            sess_conds.append("teacher_rollno = ?")
            sess_params.append(rollno)
        sess_where = " AND ".join(sess_conds)
        sess_rows = conn.execute(
            f"""SELECT DISTINCT period
                FROM attendance_sessions
                WHERE {sess_where}
                  AND period IS NOT NULL
                  AND trim(period) != ''""",
            sess_params,
        ).fetchall()
        taken_periods = {(rw["period"] or "").strip() for rw in sess_rows}

    period_headers = [str(p) for p in range(1, 10)]
    table_header_labels = {h: f"P{h}" for h in period_headers}
    if class_q and header_date and start_date_q and end_date_q and start_date_q == end_date_q:
        hdr_map: dict[str, str] = {}
        sess_params2: list = [class_q, header_date[:10]]
        teacher_clause = ""
        if r == ROLE_TEACHER and not allow_full_class_view:
            teacher_clause = " AND teacher_rollno = ?"
            sess_params2.append(rollno)
        for rw in conn.execute(
            f"""SELECT s.period,
                       COALESCE(a.subject_name, '') AS subject_name
                FROM attendance_sessions s
                LEFT JOIN attendance_subjects a ON a.subject_key = s.subject_key
                WHERE s.class_key = ?
                  AND s.session_date = ?
                  AND s.period IS NOT NULL AND trim(s.period) != ''
                  {teacher_clause}""",
            sess_params2,
        ).fetchall():
            p = str(rw["period"] or "").strip()
            nm = (rw["subject_name"] or "").strip()
            if p and nm:
                hdr_map[p] = nm
        if hdr_map:
            table_header_labels = {h: (hdr_map.get(h) or f"P{h}") for h in period_headers}

    grid_map: dict[str, dict] = {}
    if class_q and selected_class and (allow_full_class_view or subject_date_view):
        missing = parse_missing_numbers(selected_class["missing_json"] or "")
        all_rolls = build_roll_numbers(
            selected_class["roll_prefix"],
            int(selected_class["start_seq"]),
            int(selected_class["end_seq"]),
            missing,
            int(selected_class["pad_width"] or 0),
        )
        roll_names: dict[str, str] = {}
        if all_rolls:
            placeholders = ",".join("?" for _ in all_rolls)
            rows = conn.execute(
                f"SELECT rollno, name FROM users WHERE rollno IN ({placeholders})",
                all_rolls,
            ).fetchall()
            roll_names = {(rw["rollno"] or "").upper(): (rw["name"] or "") for rw in rows}
        for roll in all_rolls:
            grid_map[roll] = {
                "student_rollno": roll,
                "student_name": roll_names.get(roll.upper(), ""),
                "period_marks": {p: ("A" if p in taken_periods else "-") for p in period_headers},
            }

    for row in records:
        rl = (row["student_rollno"] or "").strip().upper()
        if not rl:
            continue
        item = grid_map.get(rl)
        if not item:
            item = {
                "student_rollno": rl,
                "student_name": (row["student_name"] or "").strip(),
                "period_marks": {p: ("A" if p in taken_periods else "-") for p in period_headers},
            }
            grid_map[rl] = item
        p = (row["period"] or "").strip()
        if p in item["period_marks"] and int(row["present_count"] or 0) > 0:
            item["period_marks"][p] = "P"

    hdr_row = ["Roll no", "Name"] + [table_header_labels[h] for h in period_headers] + ["Total present"]
    ws.append(hdr_row)
    for roll in sorted(grid_map.keys()):
        item = grid_map[roll]
        total_p = sum(1 for p in period_headers if item["period_marks"].get(p) == "P")
        ws.append(
            [item["student_rollno"], item["student_name"] or ""]
            + [item["period_marks"].get(p, "-") for p in period_headers]
            + [total_p]
        )


def _va_fill_subject_date_sheet(
    conn: sqlite3.Connection,
    ws,
    where: str,
    params: list,
    qctx: ViewAttnQ,
    r: str,
    rollno: str,
    export_limit: int,
) -> None:
    records = conn.execute(
        f"""SELECT ar.student_rollno,
                   COALESCE(u.name, '') AS student_name,
                   date(COALESCE(NULLIF(trim(ar.session_date), ''), ar.created_at)) AS period_key,
                   COUNT(*) AS present_count
            FROM attendance_records ar
            LEFT JOIN users u ON u.rollno = ar.student_rollno
            WHERE {where}
              AND ar.student_rollno IS NOT NULL
              AND trim(ar.student_rollno) != ''
            GROUP BY ar.student_rollno, u.name, period_key
            ORDER BY ar.student_rollno COLLATE NOCASE
            LIMIT {export_limit}""",
        params,
    ).fetchall()

    class_q = qctx.class_q
    subject_q = qctx.subject_q
    allow_full_class_view = qctx.allow_full_class_view
    start_date_q = qctx.start_date_q
    end_date_q = qctx.end_date_q

    selected_class = None
    taken_periods: set[str] = set()
    if class_q and (allow_full_class_view or qctx.subject_date_view):
        selected_class = conn.execute(
            """SELECT name_key, roll_prefix, start_seq, end_seq, pad_width, missing_json
               FROM attendance_classes WHERE name_key = ?""",
            (class_q,),
        ).fetchone()
        sess_conds = ["class_key = ?"]
        sess_params: list[str] = [class_q]
        if start_date_q:
            sess_conds.append("date(session_date) >= date(?)")
            sess_params.append(start_date_q)
        if end_date_q:
            sess_conds.append("date(session_date) <= date(?)")
            sess_params.append(end_date_q)
        if subject_q:
            sess_conds.append("subject_key = ?")
            sess_params.append(subject_q)
        if r == ROLE_TEACHER and not allow_full_class_view:
            sess_conds.append("teacher_rollno = ?")
            sess_params.append(rollno)
        sess_where = " AND ".join(sess_conds)
        date_rows = conn.execute(
            f"""SELECT DISTINCT session_date
                FROM attendance_sessions
                WHERE {sess_where}
                  AND session_date IS NOT NULL
                  AND trim(session_date) != ''
                ORDER BY session_date ASC""",
            sess_params,
        ).fetchall()
        taken_periods = {(rw["session_date"] or "").strip()[:10] for rw in date_rows if (rw["session_date"] or "").strip()}

    from_records = sorted(
        {str(rw["period_key"] or "").strip()[:10] for rw in records if rw["period_key"]}
    )
    period_headers = sorted(taken_periods) if taken_periods else from_records
    if not period_headers and from_records:
        period_headers = from_records
    table_header_labels = {h: h for h in period_headers}

    grid_map: dict[str, dict] = {}
    if class_q and selected_class and (allow_full_class_view or qctx.subject_date_view):
        missing = parse_missing_numbers(selected_class["missing_json"] or "")
        all_rolls = build_roll_numbers(
            selected_class["roll_prefix"],
            int(selected_class["start_seq"]),
            int(selected_class["end_seq"]),
            missing,
            int(selected_class["pad_width"] or 0),
        )
        roll_names: dict[str, str] = {}
        if all_rolls:
            placeholders = ",".join("?" for _ in all_rolls)
            rows = conn.execute(
                f"SELECT rollno, name FROM users WHERE rollno IN ({placeholders})",
                all_rolls,
            ).fetchall()
            roll_names = {(rw["rollno"] or "").upper(): (rw["name"] or "") for rw in rows}
        for roll in all_rolls:
            grid_map[roll] = {
                "student_rollno": roll,
                "student_name": roll_names.get(roll.upper(), ""),
                "period_marks": {p: ("A" if p in taken_periods else "-") for p in period_headers},
            }

    for row in records:
        rl = (row["student_rollno"] or "").strip().upper()
        if not rl:
            continue
        pk = row["period_key"]
        p = str(pk or "").strip()[:10] if pk else ""
        if not p:
            continue
        if p not in period_headers:
            continue
        item = grid_map.get(rl)
        if not item:
            item = {
                "student_rollno": rl,
                "student_name": (row["student_name"] or "").strip(),
                "period_marks": {x: ("A" if x in taken_periods else "-") for x in period_headers},
            }
            grid_map[rl] = item
        if p in item["period_marks"] and int(row["present_count"] or 0) > 0:
            item["period_marks"][p] = "P"

    hdr_row = ["Roll no", "Name"] + [table_header_labels[h] for h in period_headers] + ["Total present"]
    ws.append(hdr_row)
    for roll in sorted(grid_map.keys()):
        item = grid_map[roll]
        total_p = sum(1 for p in period_headers if item["period_marks"].get(p) == "P")
        ws.append(
            [item["student_rollno"], item["student_name"] or ""]
            + [item["period_marks"].get(p, "-") for p in period_headers]
            + [total_p]
        )


def _va_build_filtered_workbook(
    conn: sqlite3.Connection,
    qctx: ViewAttnQ,
    r: str,
    rollno: str,
    teacher_assigned_classes: set[str],
    teacher_assigned_subjects: set[str],
) -> tuple[object, str]:
    from openpyxl import Workbook

    export_limit = 100000
    conds, params = _va_base_conds(qctx, teacher_assigned_classes, teacher_assigned_subjects)
    base_where = " AND ".join(conds)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    fname_parts = ["attendance", "filtered"]

    if qctx.subject_date_view:
        ws = wb.create_sheet(title=_safe_excel_sheet_title("Subject by date"))
        _va_fill_subject_date_sheet(conn, ws, base_where, list(params), qctx, r, rollno, export_limit)
        fname_parts.append(qctx.class_q or "all")
        fname_parts.append(qctx.subject_q or "")
    elif qctx.start_date_q and qctx.end_date_q and qctx.start_date_q != qctx.end_date_q:
        day_rows = conn.execute(
            f"""SELECT DISTINCT {_ATT_AR_DAY} AS d
                FROM attendance_records ar
                WHERE {base_where}
                ORDER BY d ASC""",
            params,
        ).fetchall()
        dates = [str(rw["d"] or "").strip()[:10] for rw in day_rows if (rw["d"] or "").strip()]
        used_titles: set[str] = set()
        if not dates:
            ws = wb.create_sheet("No data")
            ws.append(["No rows in this date range for the selected filters."])
        for i, d in enumerate(dates):
            conds_d = list(conds)
            params_d = list(params)
            conds_d.append(f"{_ATT_AR_DAY} = date(?)")
            params_d.append(d)
            where_d = " AND ".join(conds_d)
            base_title = _safe_excel_sheet_title(d[:10])
            st = base_title
            n = 2
            while st in used_titles:
                st = _safe_excel_sheet_title(f"{d[:10]}_{n}")
                n += 1
            used_titles.add(st)
            ws = wb.create_sheet(title=st)
            q_day = qctx._replace(start_date_q=d, end_date_q=d)
            _va_fill_period_grid_sheet(conn, ws, where_d, params_d, q_day, r, rollno, d, export_limit)
        fname_parts.append(qctx.start_date_q or "")
        fname_parts.append(qctx.end_date_q or "")
    else:
        ws = wb.create_sheet(title=_safe_excel_sheet_title("Attendance"))
        _va_fill_period_grid_sheet(
            conn, ws, base_where, list(params), qctx, r, rollno, "", export_limit
        )
        fname_parts.append(qctx.start_date_q or qctx.end_date_q or "all")

    safe_fn = secure_filename("_".join(x for x in fname_parts if x))[:120] or "attendance_filtered"
    return wb, f"{safe_fn}.xlsx"


def _va_teacher_subject_slice_where(
    r: str,
    rollno: str,
    incharge: set[str],
    assigned_c: set[str],
    assigned_s: set[str],
) -> tuple[str, list]:
    vis_sql, vis_params = _attendance_visibility_clause(r, rollno, incharge, assigned_c, assigned_s)
    conds = [f"({vis_sql})"]
    params: list = list(vis_params)
    if r == ROLE_TEACHER and assigned_c and assigned_s:
        phc = ",".join("?" for _ in sorted(assigned_c))
        phs = ",".join("?" for _ in sorted(assigned_s))
        conds.append(f"ar.class_key IN ({phc}) AND ar.subject_key IN ({phs})")
        params.extend(sorted(assigned_c))
        params.extend(sorted(assigned_s))
    elif r == ROLE_TEACHER:
        conds.append("1 = 0")
    return " AND ".join(conds), params


def _entire_pivoted_scope_class(
    where: str,
    params: list,
    class_key_arg: str | None,
    *,
    r: str,
    assigned_c: set[str],
    incharge: set[str],
) -> tuple[str, list] | None:
    """Narrow entire-file pivoted export to one class. Teachers must have that class. Returns None if forbidden."""
    ck = (class_key_arg or "").strip()
    if not ck:
        return where, params
    if r == ROLE_TEACHER and ck not in assigned_c and ck not in incharge:
        return None
    return f"({where}) AND ar.class_key = ?", list(params) + [ck]


@app.route("/view-attendance/download")
def view_attendance_download():
    redir, r, rollno, password_only_admin, incharge, assigned_c, assigned_s = _view_attendance_auth()
    if redir is not None:
        return redir
    mode = (request.args.get("mode") or "filtered").strip().lower()
    if mode not in ("filtered", "entire"):
        return jsonify({"ok": False, "error": "Invalid mode"}), 400

    with get_db() as conn:
        if mode == "entire":
            if r == ROLE_STUDENT:
                return jsonify({"ok": False, "error": "Not available for students"}), 403
            class_key_arg = (request.args.get("class_key") or "").strip()
            want_roster_workbook = (request.args.get("format") or "").strip().lower() == "roster"
            if r == ROLE_ADMIN:
                if class_key_arg and want_roster_workbook:
                    abs_path = _safe_class_workbook_abs_path(conn, class_key_arg)
                    if not abs_path:
                        return jsonify({"ok": False, "error": "No workbook for this class"}), 404
                    return send_file(
                        abs_path,
                        as_attachment=True,
                        download_name=os.path.basename(abs_path),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                where, params = "1=1", []
                scoped = _entire_pivoted_scope_class(
                    where, params, class_key_arg, r=r, assigned_c=assigned_c, incharge=incharge
                )
                if scoped is None:
                    return jsonify({"ok": False, "error": "Invalid class scope"}), 400
                where, params = scoped
                from openpyxl import Workbook

                wb = Workbook()
                wb.remove(wb.active)
                _export_entire_attendance_subject_date_columns(conn, where, params, wb)
                bio = _workbook_bytes_xlsx(wb)
                dl_name = (
                    secure_filename(f"attendance_{class_key_arg}.xlsx")
                    if class_key_arg
                    else secure_filename("attendance_all_classes.xlsx")
                )
                return send_file(
                    bio,
                    as_attachment=True,
                    download_name=dl_name or "attendance_export.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            # teacher
            subject_key_arg = (request.args.get("subject_key") or "").strip()
            if class_key_arg and class_key_arg in incharge:
                if subject_key_arg:
                    vis_sql, vis_params = _attendance_visibility_clause(
                        r, rollno, incharge, assigned_c, assigned_s
                    )
                    where_sub = f"({vis_sql}) AND ar.class_key = ? AND ar.subject_key = ?"
                    params_sub = list(vis_params) + [class_key_arg, subject_key_arg]
                    from openpyxl import Workbook

                    wb = Workbook()
                    wb.remove(wb.active)
                    _export_entire_attendance_subject_date_columns(conn, where_sub, params_sub, wb)
                    bio = _workbook_bytes_xlsx(wb)
                    dl = secure_filename(f"attendance_{class_key_arg}_{subject_key_arg}.xlsx")[:120]
                    return send_file(
                        bio,
                        as_attachment=True,
                        download_name=dl or "attendance_subject.xlsx",
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                abs_path = _safe_class_workbook_abs_path(conn, class_key_arg)
                if not abs_path:
                    return jsonify({"ok": False, "error": "No workbook for this class"}), 404
                return send_file(
                    abs_path,
                    as_attachment=True,
                    download_name=os.path.basename(abs_path),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            if not assigned_c or not assigned_s:
                msg = (
                    "Choose your incharge class in filters, then use Download entire file for the class workbook."
                    if incharge
                    else "No subject assignments to export."
                )
                return jsonify({"ok": False, "error": msg}), 400
            where, params = _va_teacher_subject_slice_where(r, rollno, incharge, assigned_c, assigned_s)
            scoped = _entire_pivoted_scope_class(
                where, params, class_key_arg, r=r, assigned_c=assigned_c, incharge=incharge
            )
            if scoped is None:
                return jsonify({"ok": False, "error": "You are not assigned to the selected class."}), 403
            where, params = scoped
            from openpyxl import Workbook

            wb = Workbook()
            wb.remove(wb.active)
            _export_entire_attendance_subject_date_columns(conn, where, params, wb)
            bio = _workbook_bytes_xlsx(wb)
            dl_name = (
                secure_filename(f"attendance_{class_key_arg}_my_subjects.xlsx")
                if class_key_arg
                else secure_filename("attendance_my_subjects.xlsx")
            )
            return send_file(
                bio,
                as_attachment=True,
                download_name=dl_name or "attendance_my_subjects.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        qctx = _view_attendance_parse_qctx(r, rollno, incharge, assigned_c, assigned_s, request.args)
        wb, filename = _va_build_filtered_workbook(conn, qctx, r, rollno, assigned_c, assigned_s)
        bio = _workbook_bytes_xlsx(wb)
        return send_file(
            bio,
            as_attachment=True,
            download_name=secure_filename(filename),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


@app.route("/view-attendance")
def view_attendance():
    """Students see own rows; teachers see their sessions; DB admins & password admins see all."""
    password_only_admin = bool(session.get("admin_ok")) and not bool(session.get("user_id"))
    if not session.get("user_id") and not password_only_admin:
        return redirect(url_for("login"))
    if password_only_admin:
        r = ROLE_ADMIN
        rollno = ""
    else:
        r = session_role()
        rollno = session.get("rollno") or ""
        if r not in (ROLE_STUDENT, ROLE_TEACHER, ROLE_ADMIN):
            return redirect_home_for_session()
    teacher_incharge_classes: set[str] = set()
    teacher_assigned_classes: set[str] = set()
    teacher_assigned_subjects: set[str] = set()
    if r == ROLE_TEACHER and rollno:
        with get_db() as conn:
            teacher_incharge_classes, teacher_assigned_classes, teacher_assigned_subjects = _teacher_assignments(conn, rollno)
    page_q = (request.args.get("page") or "1").strip()
    try:
        page = max(1, int(page_q))
    except ValueError:
        page = 1
    qctx = _view_attendance_parse_qctx(
        r, rollno, teacher_incharge_classes, teacher_assigned_classes, teacher_assigned_subjects, request.args
    )
    start_date_q = qctx.start_date_q
    end_date_q = qctx.end_date_q
    class_q = qctx.class_q
    subject_q = qctx.subject_q
    teacher_selected_incharge = qctx.teacher_selected_incharge
    allow_full_class_view = qctx.allow_full_class_view
    subject_date_view = qctx.subject_date_view
    non_incharge_scope_only = qctx.non_incharge_scope_only
    vis_sql = qctx.vis_sql
    vis_params = list(qctx.vis_params)
    conds, params = _va_base_conds(qctx, teacher_assigned_classes, teacher_assigned_subjects)
    session_day_sql = _ATT_AR_DAY
    where = " AND ".join(conds)

    taken_periods: set[str] = set()
    editable_periods: set[str] = set()
    can_edit = False
    edit_block_reason = "Select a class and choose the same start/end date to edit attendance."
    edit_session_date = ""
    page_session_date = ""
    pagination = None
    with get_db() as conn:
        # Multi-day retrieval pagination: one attendance date per page.
        if start_date_q and end_date_q and start_date_q != end_date_q and not subject_date_view:
            sess_conds = ["date(session_date) >= date(?)", "date(session_date) <= date(?)"]
            sess_params: list[str] = [start_date_q, end_date_q]
            if class_q:
                sess_conds.append("class_key = ?")
                sess_params.append(class_q)
            if subject_q:
                sess_conds.append("subject_key = ?")
                sess_params.append(subject_q)
            if r == ROLE_TEACHER:
                teacher_scopes = ["teacher_rollno = ?"]
                teacher_scope_params: list[str] = [rollno]
                if teacher_incharge_classes:
                    ph = ",".join("?" for _ in sorted(teacher_incharge_classes))
                    teacher_scopes.append(f"class_key IN ({ph})")
                    teacher_scope_params.extend(sorted(teacher_incharge_classes))
                if teacher_assigned_classes and teacher_assigned_subjects:
                    phc = ",".join("?" for _ in sorted(teacher_assigned_classes))
                    phs = ",".join("?" for _ in sorted(teacher_assigned_subjects))
                    teacher_scopes.append(f"(class_key IN ({phc}) AND subject_key IN ({phs}))")
                    teacher_scope_params.extend(sorted(teacher_assigned_classes))
                    teacher_scope_params.extend(sorted(teacher_assigned_subjects))
                sess_conds.append("(" + " OR ".join(teacher_scopes) + ")")
                sess_params.extend(teacher_scope_params)
                if non_incharge_scope_only:
                    # Not incharge of selected class => must be within own assignments
                    if (
                        not class_q
                        or not subject_q
                        or class_q not in teacher_assigned_classes
                        or subject_q not in teacher_assigned_subjects
                    ):
                        sess_conds.append("1 = 0")
            sess_where = " AND ".join(sess_conds)
            day_rows = conn.execute(
                f"""SELECT DISTINCT session_date
                    FROM attendance_sessions
                    WHERE {sess_where}
                      AND session_date IS NOT NULL
                      AND trim(session_date) != ''
                    ORDER BY session_date ASC""",
                sess_params,
            ).fetchall()
            day_list = [(rw["session_date"] or "").strip() for rw in day_rows if (rw["session_date"] or "").strip()]
            if day_list:
                total_pages = len(day_list)
                if page > total_pages:
                    page = total_pages
                page_session_date = day_list[page - 1]
                conds.append(f"{session_day_sql} = date(?)")
                params.append(page_session_date)
                where = " AND ".join(conds)
                pagination = {
                    "enabled": True,
                    "page": page,
                    "total_pages": total_pages,
                    "has_prev": page > 1,
                    "has_next": page < total_pages,
                    "from_date": start_date_q,
                    "to_date": end_date_q,
                    "current_date": page_session_date,
                }
            else:
                pagination = {
                    "enabled": True,
                    "page": 1,
                    "total_pages": 1,
                    "has_prev": False,
                    "has_next": False,
                    "from_date": start_date_q,
                    "to_date": end_date_q,
                    "current_date": "",
                }

        ac_labels = {
            r["name_key"]: r["display_name"]
            for r in conn.execute("SELECT name_key, display_name FROM attendance_classes")
        }
        if subject_date_view:
            records = conn.execute(
                f"""SELECT ar.student_rollno,
                           COALESCE(u.name, '') AS student_name,
                           date(COALESCE(NULLIF(trim(ar.session_date), ''), ar.created_at)) AS period_key,
                           COUNT(*) AS present_count
                    FROM attendance_records ar
                    LEFT JOIN users u ON u.rollno = ar.student_rollno
                    WHERE {where}
                      AND ar.student_rollno IS NOT NULL
                      AND trim(ar.student_rollno) != ''
                    GROUP BY ar.student_rollno, u.name, period_key
                    ORDER BY ar.student_rollno COLLATE NOCASE
                    LIMIT 5000""",
                params,
            ).fetchall()
        else:
            records = conn.execute(
                f"""SELECT ar.student_rollno,
                           COALESCE(u.name, '') AS student_name,
                           ar.period,
                           COUNT(*) AS present_count
                    FROM attendance_records ar
                    LEFT JOIN users u ON u.rollno = ar.student_rollno
                    WHERE {where}
                      AND ar.student_rollno IS NOT NULL
                      AND trim(ar.student_rollno) != ''
                    GROUP BY ar.student_rollno, u.name, ar.period
                    ORDER BY ar.student_rollno COLLATE NOCASE
                    LIMIT 5000""",
                params,
            ).fetchall()
        selected_class = None
        if class_q and (allow_full_class_view or subject_date_view):
            selected_class = conn.execute(
                """SELECT name_key, roll_prefix, start_seq, end_seq, pad_width, missing_json
                   FROM attendance_classes WHERE name_key = ?""",
                (class_q,),
            ).fetchone()
            sess_conds = ["class_key = ?"]
            sess_params: list[str] = [class_q]
            if start_date_q:
                sess_conds.append("date(session_date) >= date(?)")
                sess_params.append(start_date_q)
            if end_date_q:
                sess_conds.append("date(session_date) <= date(?)")
                sess_params.append(end_date_q)
            if subject_q:
                sess_conds.append("subject_key = ?")
                sess_params.append(subject_q)
            # Incharge full-class view: periods must reflect all faculty sessions so absents show A, not "-".
            if r == ROLE_TEACHER and not allow_full_class_view:
                sess_conds.append("teacher_rollno = ?")
                sess_params.append(rollno)
            sess_where = " AND ".join(sess_conds)
            if subject_date_view:
                date_rows = conn.execute(
                    f"""SELECT DISTINCT session_date
                        FROM attendance_sessions
                        WHERE {sess_where}
                          AND session_date IS NOT NULL
                          AND trim(session_date) != ''
                        ORDER BY session_date ASC""",
                    sess_params,
                ).fetchall()
                taken_periods = {(rw["session_date"] or "").strip() for rw in date_rows if (rw["session_date"] or "").strip()}
                can_edit = False
                editable_periods = set()
                edit_block_reason = "Edit is disabled in subject date-wise view."
            else:
                sess_rows = conn.execute(
                    f"""SELECT DISTINCT period
                        FROM attendance_sessions
                        WHERE {sess_where}
                          AND period IS NOT NULL
                          AND trim(period) != ''""",
                    sess_params,
                ).fetchall()
                taken_periods = {(rw["period"] or "").strip() for rw in sess_rows}
                target_edit_date = page_session_date or (start_date_q if start_date_q and end_date_q and start_date_q == end_date_q else "")
                if target_edit_date:
                    can_edit = True
                    edit_block_reason = ""
                    edit_rows = conn.execute(
                        f"""SELECT DISTINCT period
                            FROM attendance_sessions
                            WHERE class_key = ?
                              AND session_date = ?
                              {"AND teacher_rollno = ?" if r == ROLE_TEACHER else ""}
                              AND period IS NOT NULL
                              AND trim(period) != ''""",
                        [class_q, target_edit_date, rollno] if r == ROLE_TEACHER else [class_q, target_edit_date],
                    ).fetchall()
                    editable_periods = {(rw["period"] or "").strip() for rw in edit_rows}
                    if not editable_periods:
                        can_edit = False
                        edit_block_reason = "No attendance sessions found for this class on selected date."
                    else:
                        edit_session_date = target_edit_date

        class_rows = conn.execute(
            f"""SELECT DISTINCT ar.class_key FROM attendance_records ar
                WHERE ({vis_sql})
                  AND ar.class_key IS NOT NULL AND trim(ar.class_key) != ''
                ORDER BY ar.class_key COLLATE NOCASE""",
            vis_params,
        ).fetchall()

    class_sem_filter: int | None = None
    if class_q:
        with get_db() as conn:
            class_sem_filter = _class_semester_no(conn, class_q)

    def _subject_option_rows_for_view() -> list[dict]:
        with get_db() as conn:
            sem = class_sem_filter if class_q else None
            return _attendance_subject_options(conn, sem)

    if r == ROLE_TEACHER:
        teacher_visible_classes = sorted(teacher_incharge_classes | teacher_assigned_classes)
        class_options = [
            {
                "value": ck,
                "label": ac_labels.get(ck) or CLASS_LABELS.get(ck, ck),
            }
            for ck in teacher_visible_classes
        ]
        if class_q and class_q not in {c["value"] for c in class_options}:
            class_options.insert(
                0,
                {"value": class_q, "label": ac_labels.get(class_q) or CLASS_LABELS.get(class_q, class_q)},
            )
        cand = _subject_option_rows_for_view()
        if teacher_selected_incharge:
            subject_options = [
                {"value": s["key"], "label": f"{s['name']} (Sem {s['semester_no']})"}
                for s in cand
            ]
        else:
            subject_options = [
                {"value": s["key"], "label": f"{s['name']} (Sem {s['semester_no']})"}
                for s in cand
                if s["key"] in teacher_assigned_subjects
            ]
        if subject_q and subject_q not in {s["value"] for s in subject_options}:
            subject_options.insert(0, {"value": subject_q, "label": SUBJECT_LABELS.get(subject_q, subject_q)})
    else:
        class_options = [
            {"value": k, "label": v}
            for k, v in sorted(ac_labels.items(), key=lambda x: (x[1] or "").lower())
        ]
        seen_vals = {c["value"] for c in class_options}
        for row in class_rows:
            k = row[0]
            if k and k not in seen_vals:
                class_options.append({"value": k, "label": CLASS_LABELS.get(k, k)})
                seen_vals.add(k)
        if class_q and class_q not in seen_vals:
            class_options.insert(
                0,
                {"value": class_q, "label": ac_labels.get(class_q) or CLASS_LABELS.get(class_q, class_q)},
            )
        cand = _subject_option_rows_for_view()
        subject_options = [
            {"value": s["key"], "label": f"{s['name']} (Sem {s['semester_no']})"}
            for s in cand
        ]
        if subject_q and subject_q not in {s["value"] for s in subject_options}:
            subject_options.insert(0, {"value": subject_q, "label": SUBJECT_LABELS.get(subject_q, subject_q)})

    period_headers = sorted(taken_periods) if subject_date_view else [str(p) for p in range(1, 10)]
    table_header_labels = {h: h for h in period_headers} if subject_date_view else {h: f"P{h}" for h in period_headers}
    # In full (period-wise) attendance view, show subject names as column headers
    # when a single date is selected (start == end) and class is selected.
    if not subject_date_view and class_q and start_date_q and end_date_q and start_date_q == end_date_q:
        hdr_map: dict[str, str] = {}
        with get_db() as conn:
            sess_params: list = [class_q, start_date_q]
            teacher_clause = ""
            # Non-incharge teachers should only see their own session subjects.
            if r == ROLE_TEACHER and not allow_full_class_view:
                teacher_clause = " AND teacher_rollno = ?"
                sess_params.append(rollno)
            rows = conn.execute(
                f"""SELECT s.period,
                           COALESCE(a.subject_name, '') AS subject_name
                    FROM attendance_sessions s
                    LEFT JOIN attendance_subjects a ON a.subject_key = s.subject_key
                    WHERE s.class_key = ?
                      AND s.session_date = ?
                      AND s.period IS NOT NULL AND trim(s.period) != ''
                      {teacher_clause}""",
                sess_params,
            ).fetchall()
        for rw in rows:
            p = str(rw["period"] or "").strip()
            nm = (rw["subject_name"] or "").strip()
            if p and nm:
                hdr_map[p] = nm
        if hdr_map:
            table_header_labels = {h: (hdr_map.get(h) or f"P{h}") for h in period_headers}
    grid_map: dict[str, dict] = {}

    if class_q and selected_class and (allow_full_class_view or subject_date_view):
        missing = parse_missing_numbers(selected_class["missing_json"] or "")
        all_rolls = build_roll_numbers(
            selected_class["roll_prefix"],
            int(selected_class["start_seq"]),
            int(selected_class["end_seq"]),
            missing,
            int(selected_class["pad_width"] or 0),
        )
        roll_names: dict[str, str] = {}
        if all_rolls:
            with get_db() as conn:
                placeholders = ",".join("?" for _ in all_rolls)
                rows = conn.execute(
                    f"SELECT rollno, name FROM users WHERE rollno IN ({placeholders})",
                    all_rolls,
                ).fetchall()
            roll_names = {(rw["rollno"] or "").upper(): (rw["name"] or "") for rw in rows}
        for roll in all_rolls:
            grid_map[roll] = {
                "student_rollno": roll,
                "student_name": roll_names.get(roll.upper(), ""),
                "period_marks": {p: ("A" if p in taken_periods else "-") for p in period_headers},
            }

    for row in records:
        roll = (row["student_rollno"] or "").strip().upper()
        if not roll:
            continue
        item = grid_map.get(roll)
        if not item:
            item = {
                "student_rollno": roll,
                "student_name": (row["student_name"] or "").strip(),
                "period_marks": {p: ("A" if p in taken_periods else "-") for p in period_headers},
            }
            grid_map[roll] = item
        p_val = row["period_key"] if subject_date_view else row["period"]
        p = (p_val or "").strip()
        if p in item["period_marks"] and int(row["present_count"] or 0) > 0:
            item["period_marks"][p] = "P"

    rec_list = []
    for roll in sorted(grid_map.keys()):
        item = grid_map[roll]
        item["total_present"] = sum(1 for p in period_headers if item["period_marks"][p] == "P")
        rec_list.append(item)

    unique_students = len(rec_list) if r in (ROLE_TEACHER, ROLE_ADMIN) else None

    if password_only_admin or r == ROLE_ADMIN:
        dashboard_href = url_for("admin_dashboard")
    elif r == ROLE_TEACHER:
        dashboard_href = url_for("teacher_dashboard")
    else:
        dashboard_href = url_for("student_dashboard")

    return render_template(
        "view_attendance.html",
        records=rec_list,
        period_headers=period_headers,
        table_header_labels=table_header_labels,
        class_options=class_options,
        subject_options=subject_options,
        filters={"start_date": start_date_q, "end_date": end_date_q, "class_key": class_q, "subject_key": subject_q},
        pagination=pagination,
        can_edit=can_edit,
        edit_block_reason=edit_block_reason,
        edit_session_date=edit_session_date,
        editable_periods=sorted(editable_periods),
        dashboard_href=dashboard_href,
        role=r,
        subject_date_view=subject_date_view,
        total_count=len(rec_list),
        unique_students=unique_students,
        password_only_admin=password_only_admin,
    )


@app.route("/assignments-list")
@student_or_teacher_required
def assignments_list():
    rollno = session.get("rollno")
    user = get_user_by_rollno(rollno) if rollno else None
    if not user:
        return redirect(url_for("logout"))
    return render_template("assignments_list.html", name=user["name"], rollno=user["rollno"])


@app.route("/add-credits", methods=["GET", "POST"])
@student_only_required
def add_credits():
    rollno = session.get("rollno")
    if not rollno:
        return redirect(url_for("login"))
    user = get_user_by_rollno(rollno)
    if not user:
        return redirect(url_for("logout"))
    error = None
    success = request.args.get("ok") == "1"
    if request.method == "POST":
        raw_amt = (request.form.get("amount") or "").strip()
        try:
            amount = int(raw_amt)
        except ValueError:
            amount = 0
        if amount < 1:
            error = "Enter a valid amount (at least 1 PTS)."
        file = request.files.get("screenshot")
        if not error and (not file or not file.filename):
            error = "Payment screenshot is required."
        if not error:
            ext = (os.path.splitext(file.filename)[1] or "").lower().lstrip(".")
            if ext not in ALLOWED_PHOTO_EXT:
                error = "Screenshot must be PNG, JPG, JPEG, GIF, or WEBP."
        if not error:
            safe_roll = "".join(c for c in rollno if c.isalnum() or c in "._-") or "user"
            filename = f"{secrets.token_hex(8)}_{safe_roll}.{ext}"
            filepath = os.path.join(UPLOAD_DIR_PAYMENTS, secure_filename(filename))
            if not filepath.startswith(os.path.abspath(UPLOAD_DIR_PAYMENTS)):
                error = "Invalid file path."
            else:
                file.save(filepath)
                rel = os.path.join("uploads", "payments", os.path.basename(filepath)).replace("\\", "/")
                with get_db() as conn:
                    conn.execute(
                        """INSERT INTO credit_requests (rollno, amount, screenshot_path, status)
                           VALUES (?, ?, ?, 'pending')""",
                        (rollno, amount, rel),
                    )
                    conn.commit()
                return redirect(url_for("add_credits", ok=1))
    with get_db() as conn:
        my_requests = conn.execute(
            """SELECT id, amount, status, created_at, reviewed_at
               FROM credit_requests WHERE rollno = ? ORDER BY id DESC LIMIT 20""",
            (rollno,),
        ).fetchall()
    return render_template(
        "add_credits.html",
        name=user["name"],
        rollno=rollno,
        wallet_balance=int(user["wallet_balance"] or 0),
        error=error,
        success=success,
        my_requests=my_requests,
    )


@app.route("/pay-fees")
@student_or_teacher_required
def pay_fees():
    user = get_user_by_rollno(session.get("rollno")) if session.get("rollno") else None
    return render_template("placeholder.html", title="Pay Fees", message="Fee payment portal coming soon.", name=user["name"] if user else None)


@app.route("/canteen")
@student_or_teacher_required
def canteen():
    user = get_user_by_rollno(session.get("rollno")) if session.get("rollno") else None
    return render_template("placeholder.html", title="Canteen", message="Canteen scanner coming soon.", name=user["name"] if user else None)


@app.route("/assignment-details")
@student_or_teacher_required
def assignment_details():
    rollno = session.get("rollno")
    user = get_user_by_rollno(rollno) if rollno else None
    if not user:
        return redirect(url_for("logout"))
    return render_template("assignment_details.html", name=user["name"], rollno=user["rollno"])


if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
